# -*- coding: utf-8 -*-
"""
Unit Tests fuer Export Service.

Testet CSV- und Excel-Export-Funktionen fuer deutsche Geschaeftsdokumente.
"""

import io
from datetime import date
from decimal import Decimal
from typing import Any, Dict, List
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.services.export_service import (
    INVOICE_COLUMNS,
    ORDER_COLUMNS,
    CONTRACT_COLUMNS,
    export_invoices_csv,
    export_invoices_excel,
    export_orders_csv,
    export_orders_excel,
    export_contracts_csv,
    export_contracts_excel,
    export_all_excel,
    _extract_invoice_row,
    _extract_order_row,
    _extract_contract_row,
    _format_value,
    _format_decimal_german,
)


# =============================================================================
# TEST FIXTURES
# =============================================================================

@pytest.fixture
def sample_invoice_document() -> Dict[str, Any]:
    """Beispiel-Rechnungsdokument fuer Tests."""
    return {
        "id": uuid4(),
        "filename": "rechnung_2024_001.pdf",
        "extracted_data": {
            "classification": {
                "document_type": "invoice",
                "confidence": 0.95,
            },
            "invoice": {
                "invoice_number": "RE-2024-001",
                "invoice_date": "2024-01-15",
                "due_date": "2024-02-15",
                "customer_number": "K-12345",
                "order_number": "BE-2024-100",
                "net_amount": Decimal("1000.00"),
                "vat_rate": Decimal("19.0"),
                "vat_amount": Decimal("190.00"),
                "gross_amount": Decimal("1190.00"),
                "currency": "EUR",
                "discount_percent": Decimal("2.0"),
                "discount_due_date": "2024-01-25",
                "payment_days": 30,
                "sender_vat_id": "DE123456789",
                "sender": {
                    "company": "Muster GmbH",
                    "street": "Musterstraße 123",
                    "zip_code": "12345",
                    "city": "Berlin",
                },
                "recipient": {
                    "company": "Kunde AG",
                    "street": "Kundenweg 42",
                    "zip_code": "54321",
                    "city": "München",
                },
                "sender_bank": {
                    "iban": "DE89370400440532013000",
                    "bic": "COBADEFFXXX",
                    "bank_name": "Commerzbank",
                },
            },
        },
    }


@pytest.fixture
def sample_order_document() -> Dict[str, Any]:
    """Beispiel-Bestellungsdokument fuer Tests."""
    return {
        "id": uuid4(),
        "filename": "bestellung_2024_050.pdf",
        "extracted_data": {
            "classification": {
                "document_type": "order",
                "confidence": 0.92,
            },
            "order": {
                "order_number": "BE-2024-050",
                "order_date": "2024-01-10",
                "delivery_date": "2024-01-20",
                "total_amount": Decimal("5000.00"),
                "currency": "EUR",
                "orderer": {
                    "company": "Besteller GmbH",
                    "street": "Bestellerstr. 1",
                    "city": "Hamburg",
                },
                "supplier": {
                    "company": "Lieferant AG",
                    "street": "Lieferweg 99",
                    "city": "Frankfurt",
                },
            },
        },
    }


@pytest.fixture
def sample_contract_document() -> Dict[str, Any]:
    """Beispiel-Vertragsdokument fuer Tests."""
    return {
        "id": uuid4(),
        "filename": "vertrag_2024_010.pdf",
        "extracted_data": {
            "classification": {
                "document_type": "contract",
                "confidence": 0.88,
            },
            "contract": {
                "contract_number": "VT-2024-010",
                "contract_date": "2024-01-01",
                "start_date": "2024-02-01",
                "end_date": "2025-01-31",
                "duration": "12 Monate",
                "notice_period": "3 Monate",
                "contract_value": Decimal("24000.00"),
                "monthly_value": Decimal("2000.00"),
                "contract_type": "Dienstleistungsvertrag",
                "party_a": {
                    "company": "Auftraggeber GmbH",
                    "city": "Koeln",
                },
                "party_b": {
                    "company": "Auftragnehmer AG",
                    "city": "Duesseldorf",
                },
            },
        },
    }


@pytest.fixture
def empty_document() -> Dict[str, Any]:
    """Dokument ohne extrahierte Daten."""
    return {
        "id": uuid4(),
        "filename": "leer.pdf",
        "extracted_data": None,
    }


# =============================================================================
# HELPER FUNCTION TESTS
# =============================================================================

class TestFormatValue:
    """Tests fuer _format_value Hilfsfunktion."""

    def test_format_none(self):
        """None wird zu leerem String."""
        assert _format_value(None) == ""

    def test_format_string(self):
        """String bleibt unveraendert."""
        assert _format_value("Test") == "Test"

    def test_format_date(self):
        """Date wird zu ISO-String."""
        d = date(2024, 1, 15)
        assert _format_value(d) == "2024-01-15"

    def test_format_decimal(self):
        """Decimal wird zu String."""
        d = Decimal("1234.56")
        assert _format_value(d) == "1234.56"

    def test_format_float(self):
        """Float wird mit 2 Dezimalstellen formatiert."""
        assert _format_value(1234.5) == "1234.50"
        assert _format_value(1234.567) == "1234.57"

    def test_format_integer(self):
        """Integer wird zu String."""
        assert _format_value(42) == "42"


class TestFormatDecimalGerman:
    """Tests fuer _format_decimal_german Hilfsfunktion."""

    def test_format_none_returns_none(self):
        """None wird zu None."""
        assert _format_decimal_german(None) is None

    def test_format_decimal(self):
        """Decimal wird zu float."""
        assert _format_decimal_german(Decimal("1234.56")) == 1234.56

    def test_format_string_number(self):
        """Numerischer String wird zu float."""
        assert _format_decimal_german("1234.56") == 1234.56

    def test_format_invalid_string_returns_none(self):
        """Nicht-numerischer String wird zu None."""
        assert _format_decimal_german("nicht eine zahl") is None


# =============================================================================
# ROW EXTRACTION TESTS
# =============================================================================

class TestExtractInvoiceRow:
    """Tests fuer _extract_invoice_row."""

    def test_extract_complete_invoice(self, sample_invoice_document):
        """Vollstaendige Rechnung wird korrekt extrahiert."""
        row = _extract_invoice_row(sample_invoice_document)

        assert row["invoice_number"] == "RE-2024-001"
        assert row["invoice_date"] == "2024-01-15"
        assert row["sender_company"] == "Muster GmbH"
        assert row["sender_iban"] == "DE89370400440532013000"
        assert row["recipient_company"] == "Kunde AG"
        assert row["net_amount"] == Decimal("1000.00")
        assert row["gross_amount"] == Decimal("1190.00")
        assert row["confidence"] == 0.95

    def test_extract_empty_document(self, empty_document):
        """Leeres Dokument gibt None-Werte zurueck."""
        row = _extract_invoice_row(empty_document)

        assert row["invoice_number"] is None
        assert row["sender_company"] is None
        assert row["net_amount"] is None
        # Confidence hat Default-Wert
        assert row["confidence"] == 0.0

    def test_extract_partial_invoice(self):
        """Teilweise gefuellte Rechnung wird korrekt verarbeitet."""
        doc = {
            "id": uuid4(),
            "filename": "partial.pdf",
            "extracted_data": {
                "invoice": {
                    "invoice_number": "RE-PARTIAL",
                    # Fehlende Felder
                },
            },
        }
        row = _extract_invoice_row(doc)

        assert row["invoice_number"] == "RE-PARTIAL"
        assert row["sender_company"] is None


class TestExtractOrderRow:
    """Tests fuer _extract_order_row."""

    def test_extract_complete_order(self, sample_order_document):
        """Vollstaendige Bestellung wird korrekt extrahiert."""
        row = _extract_order_row(sample_order_document)

        assert row["order_number"] == "BE-2024-050"
        assert row["orderer_company"] == "Besteller GmbH"
        assert row["supplier_company"] == "Lieferant AG"
        assert row["total_amount"] == Decimal("5000.00")


class TestExtractContractRow:
    """Tests fuer _extract_contract_row."""

    def test_extract_complete_contract(self, sample_contract_document):
        """Vollstaendiger Vertrag wird korrekt extrahiert."""
        row = _extract_contract_row(sample_contract_document)

        assert row["contract_number"] == "VT-2024-010"
        assert row["party_a_company"] == "Auftraggeber GmbH"
        assert row["party_b_company"] == "Auftragnehmer AG"
        assert row["contract_value"] == Decimal("24000.00")
        assert row["contract_type"] == "Dienstleistungsvertrag"


# =============================================================================
# CSV EXPORT TESTS
# =============================================================================

class TestExportInvoicesCsv:
    """Tests fuer export_invoices_csv."""

    def test_export_single_invoice(self, sample_invoice_document):
        """Eine Rechnung wird korrekt als CSV exportiert."""
        csv_content = export_invoices_csv([sample_invoice_document])

        # UTF-8 BOM pruefen
        assert csv_content.startswith('\ufeff')

        lines = csv_content.strip().split('\n')
        assert len(lines) == 2  # Header + 1 Zeile

        # Header pruefen (deutsche Spaltennamen)
        header = lines[0].replace('\ufeff', '')
        assert "Rechnungsnummer" in header
        assert "Absender (Firma)" in header

        # Daten pruefen
        data_line = lines[1]
        assert "RE-2024-001" in data_line
        assert "Muster GmbH" in data_line

    def test_export_multiple_invoices(self, sample_invoice_document):
        """Mehrere Rechnungen werden korrekt exportiert."""
        docs = [sample_invoice_document, sample_invoice_document]
        csv_content = export_invoices_csv(docs)

        lines = csv_content.strip().split('\n')
        assert len(lines) == 3  # Header + 2 Zeilen

    def test_export_empty_list(self):
        """Leere Liste ergibt nur Header."""
        csv_content = export_invoices_csv([])

        lines = csv_content.strip().split('\n')
        assert len(lines) == 1  # Nur Header

    def test_csv_uses_semicolon_delimiter(self, sample_invoice_document):
        """CSV verwendet Semikolon als Trennzeichen (fuer deutsches Excel)."""
        csv_content = export_invoices_csv([sample_invoice_document])

        # Semikolon im Header pruefen
        assert ';' in csv_content


class TestExportOrdersCsv:
    """Tests fuer export_orders_csv."""

    def test_export_single_order(self, sample_order_document):
        """Eine Bestellung wird korrekt exportiert."""
        csv_content = export_orders_csv([sample_order_document])

        assert "Bestellnummer" in csv_content
        assert "BE-2024-050" in csv_content


class TestExportContractsCsv:
    """Tests fuer export_contracts_csv."""

    def test_export_single_contract(self, sample_contract_document):
        """Ein Vertrag wird korrekt exportiert."""
        csv_content = export_contracts_csv([sample_contract_document])

        assert "Vertragsnummer" in csv_content
        assert "VT-2024-010" in csv_content


# =============================================================================
# EXCEL EXPORT TESTS
# =============================================================================

class TestExportInvoicesExcel:
    """Tests fuer export_invoices_excel."""

    def test_export_single_invoice(self, sample_invoice_document):
        """Eine Rechnung wird als Excel exportiert."""
        excel_bytes = export_invoices_excel([sample_invoice_document])

        # Als gueltige Excel-Datei laden
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active

        # Tab-Name
        assert ws.title == "Rechnungen"

        # Header in Zeile 1
        assert ws.cell(1, 1).value == "Rechnungsnummer"

        # Daten in Zeile 2
        assert ws.cell(2, 1).value == "RE-2024-001"

    def test_excel_has_autofilter(self, sample_invoice_document):
        """Excel hat Autofilter aktiviert."""
        excel_bytes = export_invoices_excel([sample_invoice_document])
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active

        assert ws.auto_filter.ref is not None

    def test_excel_has_frozen_panes(self, sample_invoice_document):
        """Excel hat fixierte erste Zeile."""
        excel_bytes = export_invoices_excel([sample_invoice_document])
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active

        assert ws.freeze_panes == "A2"

    def test_excel_styling(self, sample_invoice_document):
        """Excel hat formatierte Header."""
        excel_bytes = export_invoices_excel([sample_invoice_document])
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active

        # Header-Zelle hat fette Schrift
        header_cell = ws.cell(1, 1)
        assert header_cell.font.bold is True

    def test_export_empty_list(self):
        """Leere Liste erzeugt Excel nur mit Header."""
        excel_bytes = export_invoices_excel([])
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active

        # Header vorhanden
        assert ws.cell(1, 1).value == "Rechnungsnummer"
        # Keine Datenzeilen
        assert ws.cell(2, 1).value is None


class TestExportOrdersExcel:
    """Tests fuer export_orders_excel."""

    def test_export_single_order(self, sample_order_document):
        """Eine Bestellung wird als Excel exportiert."""
        excel_bytes = export_orders_excel([sample_order_document])
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active

        assert ws.title == "Bestellungen"
        assert ws.cell(2, 1).value == "BE-2024-050"


class TestExportContractsExcel:
    """Tests fuer export_contracts_excel."""

    def test_export_single_contract(self, sample_contract_document):
        """Ein Vertrag wird als Excel exportiert."""
        excel_bytes = export_contracts_excel([sample_contract_document])
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active

        assert ws.title == "Vertraege"
        assert ws.cell(2, 1).value == "VT-2024-010"


# =============================================================================
# COMBINED EXCEL EXPORT TESTS
# =============================================================================

class TestExportAllExcel:
    """Tests fuer export_all_excel (kombinierter Export)."""

    def test_export_all_document_types(
        self,
        sample_invoice_document,
        sample_order_document,
        sample_contract_document
    ):
        """Alle Dokumenttypen in einer Excel-Datei mit separaten Tabs."""
        excel_bytes = export_all_excel(
            invoices=[sample_invoice_document],
            orders=[sample_order_document],
            contracts=[sample_contract_document]
        )

        wb = load_workbook(io.BytesIO(excel_bytes))

        # Drei Tabs vorhanden
        assert len(wb.sheetnames) == 3
        assert "Rechnungen" in wb.sheetnames
        assert "Bestellungen" in wb.sheetnames
        assert "Vertraege" in wb.sheetnames

        # Daten in jedem Tab
        assert wb["Rechnungen"].cell(2, 1).value == "RE-2024-001"
        assert wb["Bestellungen"].cell(2, 1).value == "BE-2024-050"
        assert wb["Vertraege"].cell(2, 1).value == "VT-2024-010"

    def test_export_with_empty_categories(self):
        """Export funktioniert auch wenn einzelne Kategorien leer sind."""
        excel_bytes = export_all_excel(
            invoices=[],
            orders=[],
            contracts=[]
        )

        wb = load_workbook(io.BytesIO(excel_bytes))

        # Alle Tabs vorhanden
        assert len(wb.sheetnames) == 3

        # Header in jedem Tab
        assert wb["Rechnungen"].cell(1, 1).value == "Rechnungsnummer"
        assert wb["Bestellungen"].cell(1, 1).value == "Bestellnummer"
        assert wb["Vertraege"].cell(1, 1).value == "Vertragsnummer"


# =============================================================================
# EDGE CASE TESTS
# =============================================================================

class TestEdgeCases:
    """Tests fuer Randfaelle."""

    def test_document_with_umlauts(self):
        """Deutsche Umlaute werden korrekt verarbeitet."""
        doc = {
            "id": uuid4(),
            "filename": "rechnung_ümläut.pdf",
            "extracted_data": {
                "invoice": {
                    "invoice_number": "RE-ÄÖÜ-001",
                    "sender": {
                        "company": "Müller & Söhne GmbH",
                        "city": "Düsseldorf",
                    },
                },
            },
        }

        csv_content = export_invoices_csv([doc])

        assert "RE-ÄÖÜ-001" in csv_content
        assert "Müller & Söhne GmbH" in csv_content
        assert "Düsseldorf" in csv_content

    def test_document_with_special_characters(self):
        """Sonderzeichen werden korrekt verarbeitet."""
        doc = {
            "id": uuid4(),
            "filename": "test.pdf",
            "extracted_data": {
                "invoice": {
                    "invoice_number": 'RE-"Test";123',
                    "sender": {
                        "company": "Test & Co, Ltd.",
                    },
                },
            },
        }

        csv_content = export_invoices_csv([doc])

        # Sonderzeichen sollten escaped/quotiert sein
        assert "Test" in csv_content

    def test_document_with_very_large_numbers(self):
        """Grosse Zahlen werden korrekt verarbeitet."""
        doc = {
            "id": uuid4(),
            "filename": "gross.pdf",
            "extracted_data": {
                "invoice": {
                    "gross_amount": Decimal("9999999999.99"),
                },
            },
        }

        excel_bytes = export_invoices_excel([doc])
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active

        # Zahl sollte als Zahl gespeichert sein
        gross_col = None
        for col, (header, field) in enumerate(INVOICE_COLUMNS, 1):
            if field == "gross_amount":
                gross_col = col
                break

        value = ws.cell(2, gross_col).value
        assert value == 9999999999.99


class TestColumnDefinitions:
    """Tests fuer Spaltendefinitionen."""

    def test_invoice_columns_have_correct_count(self):
        """Rechnungsspalten haben korrekte Anzahl."""
        assert len(INVOICE_COLUMNS) == 23

    def test_order_columns_have_correct_count(self):
        """Bestellspalten haben korrekte Anzahl."""
        assert len(ORDER_COLUMNS) == 10

    def test_contract_columns_have_correct_count(self):
        """Vertragsspalten haben korrekte Anzahl."""
        assert len(CONTRACT_COLUMNS) == 14

    def test_column_definitions_are_tuples(self):
        """Alle Spaltendefinitionen sind (header, field) Tupel."""
        for col in INVOICE_COLUMNS + ORDER_COLUMNS + CONTRACT_COLUMNS:
            assert isinstance(col, tuple)
            assert len(col) == 2
            assert isinstance(col[0], str)  # Header
            assert isinstance(col[1], str)  # Field name
