# -*- coding: utf-8 -*-
"""
Tests fuer ExcelReportGenerator.

Testet:
- Report-Erstellung
- Sheet-Erstellung
- Chart-Erstellung
- Customer/Supplier Reports
- Edge Cases und Fehlerbehandlung
"""

import pytest
from datetime import datetime
from pathlib import Path
from io import BytesIO
from unittest.mock import MagicMock, patch
from typing import Dict, Any, List

# Imports mit Fallback
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from app.services.rag.excel_generator import (
    ExcelReportGenerator,
    get_excel_generator,
    OPENPYXL_AVAILABLE as MODULE_OPENPYXL,
)


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl nicht installiert")
class TestExcelReportGeneratorInit:
    """Tests fuer Generator-Initialisierung."""

    def test_init_with_openpyxl(self):
        """Sollte Generator initialisieren wenn openpyxl verfuegbar."""
        generator = ExcelReportGenerator()

        assert generator.HEADER_FONT is not None
        assert generator.HEADER_FILL is not None

    def test_styles_defined(self):
        """Sollte Standard-Styles definiert haben."""
        generator = ExcelReportGenerator()

        assert generator.TITLE_FONT is not None
        assert generator.SUBTITLE_FONT is not None


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl nicht installiert")
class TestCreateReport:
    """Tests fuer create_report Methode."""

    @pytest.fixture
    def generator(self):
        return ExcelReportGenerator()

    def test_create_simple_report(self, generator: ExcelReportGenerator):
        """Sollte einfachen Report erstellen."""
        data = {
            "summary": {
                "text": "Test-Zusammenfassung",
                "metrics": {"Metrik 1": 100, "Metrik 2": 200}
            },
            "sheets": {},
            "charts": []
        }

        result = generator.create_report("Test Report", data)

        assert isinstance(result, bytes)
        assert len(result) > 0

        # Verify Excel content
        wb = load_workbook(BytesIO(result))
        assert "Zusammenfassung" in wb.sheetnames

    def test_create_report_with_data_sheet(self, generator: ExcelReportGenerator):
        """Sollte Report mit Daten-Sheet erstellen."""
        data = {
            "summary": {},
            "sheets": {
                "Testdaten": {
                    "headers": ["Name", "Wert", "Status"],
                    "rows": [
                        ["Item 1", 100, "Aktiv"],
                        ["Item 2", 200, "Inaktiv"],
                    ]
                }
            },
            "charts": []
        }

        result = generator.create_report("Test Report", data)

        wb = load_workbook(BytesIO(result))
        assert "Testdaten" in wb.sheetnames

        # Pruefen ob Daten korrekt sind
        ws = wb["Testdaten"]
        assert ws["A1"].value == "Name"
        assert ws["A2"].value == "Item 1"
        assert ws["B2"].value == 100

    def test_create_report_with_currency_columns(self, generator: ExcelReportGenerator):
        """Sollte Waehrungsspalten formatieren."""
        data = {
            "summary": {},
            "sheets": {
                "Finanzen": {
                    "headers": ["Beschreibung", "Betrag"],
                    "rows": [
                        ["Rechnung 1", 1234.56],
                        ["Rechnung 2", 789.00],
                    ],
                    "currency_columns": [2]
                }
            },
            "charts": []
        }

        result = generator.create_report("Finanz-Report", data)

        wb = load_workbook(BytesIO(result))
        ws = wb["Finanzen"]
        # Waehrungsformat sollte angewendet sein
        assert ws["B2"].value == 1234.56

    def test_create_report_with_chart(self, generator: ExcelReportGenerator):
        """Sollte Report mit Chart erstellen."""
        data = {
            "summary": {},
            "sheets": {
                "Daten": {
                    "headers": ["Kategorie", "Wert"],
                    "rows": [
                        ["A", 10],
                        ["B", 20],
                        ["C", 30],
                    ]
                }
            },
            "charts": [
                {
                    "type": "bar",
                    "title": "Test Chart",
                    "sheet": "Daten",
                    "data_range": {
                        "min_col": 2,
                        "max_col": 2,
                        "min_row": 1,
                        "max_row": 4,
                        "cat_col": 1
                    },
                    "position": "D2"
                }
            ]
        }

        result = generator.create_report("Chart Report", data)

        wb = load_workbook(BytesIO(result))
        assert "Daten" in wb.sheetnames

    def test_create_report_save_to_file(self, generator: ExcelReportGenerator, tmp_path):
        """Sollte Report in Datei speichern."""
        data = {
            "summary": {"text": "Test"},
            "sheets": {},
            "charts": []
        }

        output_path = tmp_path / "test_report.xlsx"
        result = generator.create_report("Test", data, output_path=output_path)

        assert output_path.exists()
        assert output_path.read_bytes() == result

    def test_create_report_long_sheet_name(self, generator: ExcelReportGenerator):
        """Sollte lange Sheet-Namen kuerzen (max 31 Zeichen)."""
        data = {
            "summary": {},
            "sheets": {
                "Dies ist ein sehr langer Sheet-Name der gekuerzt werden muss": {
                    "headers": ["A"],
                    "rows": [["test"]]
                }
            },
            "charts": []
        }

        result = generator.create_report("Test", data)

        wb = load_workbook(BytesIO(result))
        # Alle Sheet-Namen sollten <= 31 Zeichen sein
        for name in wb.sheetnames:
            assert len(name) <= 31


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl nicht installiert")
class TestCreateCustomerReport:
    """Tests fuer create_customer_report Methode."""

    @pytest.fixture
    def generator(self):
        return ExcelReportGenerator()

    def test_create_customer_report_minimal(self, generator: ExcelReportGenerator):
        """Sollte minimalen Kundenreport erstellen."""
        result = generator.create_customer_report(
            customer_name="Test GmbH",
            summary="Dies ist eine Test-Zusammenfassung.",
            documents=[],
            metrics={"Dokumente": 0}
        )

        assert isinstance(result, bytes)

        wb = load_workbook(BytesIO(result))
        assert "Zusammenfassung" in wb.sheetnames

    def test_create_customer_report_with_documents(self, generator: ExcelReportGenerator):
        """Sollte Kundenreport mit Dokumenten erstellen."""
        documents = [
            {"date": "2024-01-15", "type": "Rechnung", "title": "RE-2024-001", "status": "Bezahlt"},
            {"date": "2024-02-20", "type": "Vertrag", "title": "VT-2024-001", "status": "Aktiv"},
            {"date": "2024-03-10", "type": "Rechnung", "title": "RE-2024-002", "status": "Offen"},
        ]

        result = generator.create_customer_report(
            customer_name="Muster AG",
            summary="Wichtiger Kunde mit mehreren Dokumenten.",
            documents=documents,
            metrics={
                "Gesamtwert": "€ 50.000",
                "Offene Rechnungen": 1
            }
        )

        wb = load_workbook(BytesIO(result))
        assert "Dokumente" in wb.sheetnames
        assert "Dokumenttypen" in wb.sheetnames

        # Pruefen ob Dokumenttypen-Verteilung korrekt ist
        ws = wb["Dokumenttypen"]
        # Header + 2 Typen (Rechnung, Vertrag)
        assert ws["A2"].value in ["Rechnung", "Vertrag"]

    def test_create_customer_report_with_pie_chart(self, generator: ExcelReportGenerator):
        """Sollte Kundenreport mit Pie-Chart erstellen."""
        documents = [
            {"type": "Rechnung"},
            {"type": "Rechnung"},
            {"type": "Vertrag"},
        ]

        result = generator.create_customer_report(
            customer_name="Test",
            summary="Test",
            documents=documents,
            metrics={}
        )

        wb = load_workbook(BytesIO(result))
        # Chart sollte im Dokumenttypen-Sheet sein
        assert "Dokumenttypen" in wb.sheetnames


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl nicht installiert")
class TestCreateSupplierReport:
    """Tests fuer create_supplier_report Methode."""

    @pytest.fixture
    def generator(self):
        return ExcelReportGenerator()

    def test_create_supplier_report_minimal(self, generator: ExcelReportGenerator):
        """Sollte minimalen Lieferantenreport erstellen."""
        result = generator.create_supplier_report(
            supplier_name="Lieferant GmbH",
            summary="Zuverlaessiger Lieferant.",
            invoices=[],
            contracts=[],
            metrics={}
        )

        assert isinstance(result, bytes)

        wb = load_workbook(BytesIO(result))
        assert "Zusammenfassung" in wb.sheetnames
        assert "Rechnungen" in wb.sheetnames
        assert "Verträge" in wb.sheetnames

    def test_create_supplier_report_with_data(self, generator: ExcelReportGenerator):
        """Sollte Lieferantenreport mit Daten erstellen."""
        invoices = [
            {
                "number": "RE-001",
                "date": "2024-01-15",
                "amount": 1500.00,
                "status": "Bezahlt",
                "due_date": "2024-02-15"
            },
        ]
        contracts = [
            {
                "number": "VT-001",
                "title": "Rahmenvertrag",
                "start_date": "2024-01-01",
                "end_date": "2024-12-31",
                "value": 50000.00
            },
        ]

        result = generator.create_supplier_report(
            supplier_name="Lieferant AG",
            summary="Wichtiger Lieferant.",
            invoices=invoices,
            contracts=contracts,
            metrics={"Umsatz 2024": "€ 50.000"}
        )

        wb = load_workbook(BytesIO(result))

        # Rechnungen pruefen
        ws_invoices = wb["Rechnungen"]
        assert ws_invoices["A1"].value == "Rechnungsnr."
        assert ws_invoices["A2"].value == "RE-001"

        # Vertraege pruefen
        ws_contracts = wb["Verträge"]
        assert ws_contracts["A1"].value == "Vertragsnr."
        assert ws_contracts["B2"].value == "Rahmenvertrag"


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl nicht installiert")
class TestChartTypes:
    """Tests fuer verschiedene Chart-Typen."""

    @pytest.fixture
    def generator(self):
        return ExcelReportGenerator()

    def _create_report_with_chart(self, generator, chart_type):
        """Hilfsmethode fuer Chart-Tests."""
        data = {
            "summary": {},
            "sheets": {
                "Daten": {
                    "headers": ["X", "Y"],
                    "rows": [[1, 10], [2, 20], [3, 30]]
                }
            },
            "charts": [{
                "type": chart_type,
                "title": f"{chart_type} Chart",
                "sheet": "Daten",
                "data_range": {
                    "min_col": 2, "max_col": 2,
                    "min_row": 1, "max_row": 4,
                    "cat_col": 1
                },
                "position": "D2"
            }]
        }
        return generator.create_report("Chart Test", data)

    def test_bar_chart(self, generator: ExcelReportGenerator):
        """Sollte Bar-Chart erstellen."""
        result = self._create_report_with_chart(generator, "bar")
        assert isinstance(result, bytes)

    def test_pie_chart(self, generator: ExcelReportGenerator):
        """Sollte Pie-Chart erstellen."""
        result = self._create_report_with_chart(generator, "pie")
        assert isinstance(result, bytes)

    def test_line_chart(self, generator: ExcelReportGenerator):
        """Sollte Line-Chart erstellen."""
        result = self._create_report_with_chart(generator, "line")
        assert isinstance(result, bytes)

    def test_unknown_chart_type_defaults_to_bar(self, generator: ExcelReportGenerator):
        """Sollte bei unbekanntem Chart-Typ Bar verwenden."""
        result = self._create_report_with_chart(generator, "unknown")
        assert isinstance(result, bytes)


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl nicht installiert")
class TestEdgeCases:
    """Tests fuer Randfaelle."""

    @pytest.fixture
    def generator(self):
        return ExcelReportGenerator()

    def test_empty_data(self, generator: ExcelReportGenerator):
        """Sollte leere Daten verarbeiten."""
        data = {
            "summary": {},
            "sheets": {},
            "charts": []
        }

        result = generator.create_report("Empty Report", data)
        assert isinstance(result, bytes)

    def test_empty_rows(self, generator: ExcelReportGenerator):
        """Sollte leere Zeilen verarbeiten."""
        data = {
            "summary": {},
            "sheets": {
                "Leer": {
                    "headers": ["A", "B"],
                    "rows": []
                }
            },
            "charts": []
        }

        result = generator.create_report("Empty Rows", data)
        wb = load_workbook(BytesIO(result))
        assert "Leer" in wb.sheetnames

    def test_chart_on_nonexistent_sheet(self, generator: ExcelReportGenerator):
        """Sollte Chart auf nicht existierendem Sheet ignorieren."""
        data = {
            "summary": {},
            "sheets": {},
            "charts": [{
                "type": "bar",
                "sheet": "NichtVorhanden",
                "data_range": {}
            }]
        }

        # Sollte keinen Fehler werfen
        result = generator.create_report("Test", data)
        assert isinstance(result, bytes)

    def test_unicode_content(self, generator: ExcelReportGenerator):
        """Sollte Unicode-Inhalte verarbeiten."""
        data = {
            "summary": {
                "text": "Zusammenfassung mit Umlauten: äöüß",
                "metrics": {"Größe": "100"}
            },
            "sheets": {
                "Übersicht": {
                    "headers": ["Müller", "Größe"],
                    "rows": [["Böhm", "Größer"]]
                }
            },
            "charts": []
        }

        result = generator.create_report("Ümläut Report", data)

        wb = load_workbook(BytesIO(result))
        assert "Übersicht" in wb.sheetnames


class TestSingleton:
    """Tests fuer Singleton-Pattern."""

    def test_get_excel_generator_singleton(self):
        """Sollte immer gleiche Instanz zurueckgeben."""
        # Reset singleton
        import app.services.rag.excel_generator as module
        module._excel_generator = None

        gen1 = get_excel_generator()
        gen2 = get_excel_generator()

        assert gen1 is gen2


class TestWithoutOpenpyxl:
    """Tests wenn openpyxl nicht verfuegbar ist."""

    def test_create_report_raises_without_openpyxl(self):
        """Sollte ImportError werfen wenn openpyxl fehlt."""
        with patch.object(
            ExcelReportGenerator, '__init__',
            return_value=None
        ):
            generator = ExcelReportGenerator.__new__(ExcelReportGenerator)

            with patch(
                'app.services.rag.excel_generator.OPENPYXL_AVAILABLE',
                False
            ):
                with pytest.raises(ImportError, match="openpyxl"):
                    generator.create_report("Test", {})
