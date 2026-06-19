# -*- coding: utf-8 -*-
"""
Invoice Tracking API Endpoints.

REST API für Rechnungsverfolgung (Risk Scoring):
- CRUD Operationen für InvoiceTracking
- Zahlungsstatus-Updates
- Mahnwesen-Integration
- Verknüpfung mit Dokumenten

Feinpoliert und durchdacht - Enterprise Risk Scoring.
"""

from typing import Optional, List
from app.api.dependencies import get_user_company_id_dep  # F-31
from app.api.dependencies import get_user_company_id  # F-31
from uuid import UUID
from datetime import datetime, timezone

import structlog
from fastapi import APIRouter, Depends, HTTPException, status, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func

from app.db.models import User, InvoiceTracking, Document, InvoiceStatus
from app.db.schemas import (
    InvoiceTrackingCreate,
    InvoiceTrackingUpdate,
    InvoiceTrackingResponse,
    InvoiceStatusEnum,
    InvoiceStatisticsResponse,
)
from app.api.dependencies import (
    get_db,
    get_current_active_user,
    get_user_company_id,  # noqa: F401 - Re-Export (B1: zentralisiert in dependencies.py)
    get_user_company_id_dep,
)
from app.workers.tasks.risk_scoring_tasks import on_invoice_updated_recalculate
from app.core.safe_errors import safe_error_detail, safe_error_log
from app.core.security_auth import build_content_disposition

logger = structlog.get_logger(__name__)

router = APIRouter(prefix="/invoices", tags=["Invoice Tracking"])


# =============================================================================
# Helper Functions - Multi-Tenant Security
# =============================================================================
# B1: Die Multi-Tenant-Helfer (get_user_company_id, _require_user_company_id,
# get_user_company_id_dep) wurden zentral nach app/api/dependencies.py verschoben
# und oben re-importiert. Endpoints nutzen get_user_company_id_dep als Dependency.


# =============================================================================
# LIST / SEARCH
# =============================================================================


@router.get(
    "",
    response_model=List[InvoiceTrackingResponse],
    summary="Rechnungen auflisten",
    description="Listet alle Rechnungsverfolgungen mit Filter-Optionen"
)
async def list_invoices(
    page: int = Query(1, ge=1, description="Seitennummer"),
    per_page: int = Query(20, ge=1, le=100, description="Einträge pro Seite"),
    status_filter: Optional[InvoiceStatusEnum] = Query(
        None, alias="status", description="Nach Status filtern"
    ),
    overdue_only: bool = Query(False, description="Nur überfällige Rechnungen"),
    document_id: Optional[UUID] = Query(None, description="Nach Dokument filtern"),
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
    company_id: UUID = Depends(get_user_company_id_dep),
) -> List[InvoiceTrackingResponse]:
    """
    Listet alle Rechnungsverfolgungen auf.

    **Filter:**
    - **status**: open, sent, paid, overdue, dunning, cancelled, partial
    - **overdue_only**: Nur überfällige Rechnungen
    - **document_id**: Nur Rechnungen eines bestimmten Dokuments
    """
    # SECURITY: Multi-Tenant RLS - nur Rechnungen zu Dokumenten des aktuellen Users
    query = (
        select(InvoiceTracking)
        .join(Document, InvoiceTracking.document_id == Document.id)
        .where(
            and_(
                InvoiceTracking.deleted_at.is_(None),
                Document.company_id == company_id,
            )
        )
    )

    if status_filter:
        query = query.where(InvoiceTracking.status == status_filter.value)

    if overdue_only:
        now = datetime.now(timezone.utc)
        query = query.where(
            and_(
                InvoiceTracking.status.notin_(["paid", "cancelled"]),
                InvoiceTracking.due_date < now,
            )
        )

    if document_id:
        query = query.where(InvoiceTracking.document_id == document_id)

    # Pagination
    offset = (page - 1) * per_page
    query = query.order_by(InvoiceTracking.created_at.desc())
    query = query.offset(offset).limit(per_page)

    result = await db.execute(query)
    invoices = result.scalars().all()

    # Compute is_overdue and days_overdue for each invoice
    response_list = []
    now = datetime.now(timezone.utc)
    for inv in invoices:
        resp = InvoiceTrackingResponse.model_validate(inv)
        # Compute overdue status
        if inv.status not in ("paid", "cancelled") and inv.due_date:
            due_date = inv.due_date if inv.due_date.tzinfo else inv.due_date.replace(tzinfo=timezone.utc)
            if now > due_date:
                resp.is_overdue = True
                resp.days_overdue = (now - due_date).days
        response_list.append(resp)

    return response_list


# =============================================================================
# STATISTICS (MUSS vor /{invoice_id} stehen - FastAPI Route-Reihenfolge!)
# =============================================================================


@router.get(
    "/statistics/summary",
    response_model=InvoiceStatisticsResponse,
    summary="Rechnungsstatistiken abrufen",
    description="Liefert aggregierte Statistiken über alle Rechnungen"
)
async def get_invoice_statistics(
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
    company_id: UUID = Depends(get_user_company_id_dep),
) -> InvoiceStatisticsResponse:
    """
    Liefert aggregierte Rechnungsstatistiken.

    **Response:**
    - Anzahl nach Status
    - Gesamtbetraege
    - Durchschnittliche Zahlungsverzögerung
    """
    # SECURITY: Multi-Tenant RLS - nur Statistiken für eigene Rechnungen
    base_query = (
        select(
            func.count(InvoiceTracking.id).label("total"),
            func.sum(InvoiceTracking.amount).label("total_amount"),
        )
        .select_from(InvoiceTracking)
        .join(Document, InvoiceTracking.document_id == Document.id)
        .where(
            and_(
                InvoiceTracking.deleted_at.is_(None),
                Document.company_id == company_id,
            )
        )
    )

    result = await db.execute(base_query)
    stats = result.one()

    # Status-Verteilung (SECURITY: Multi-Tenant RLS)
    status_query = (
        select(
            InvoiceTracking.status,
            func.count(InvoiceTracking.id).label("count"),
            func.sum(InvoiceTracking.amount).label("amount"),
        )
        .select_from(InvoiceTracking)
        .join(Document, InvoiceTracking.document_id == Document.id)
        .where(
            and_(
                InvoiceTracking.deleted_at.is_(None),
                Document.company_id == company_id,
            )
        )
        .group_by(InvoiceTracking.status)
    )

    status_result = await db.execute(status_query)
    status_distribution = {
        row.status: {"count": row.count, "amount": row.amount or 0.0}
        for row in status_result
    }

    # Überfällige Rechnungen (SECURITY: Multi-Tenant RLS)
    now = datetime.now(timezone.utc)
    overdue_query = (
        select(
            func.count(InvoiceTracking.id).label("count"),
            func.sum(InvoiceTracking.amount).label("amount"),
        )
        .select_from(InvoiceTracking)
        .join(Document, InvoiceTracking.document_id == Document.id)
        .where(
            and_(
                InvoiceTracking.deleted_at.is_(None),
                Document.company_id == company_id,
                InvoiceTracking.status.notin_(["paid", "cancelled"]),
                InvoiceTracking.due_date < now,
            )
        )
    )

    overdue_result = await db.execute(overdue_query)
    overdue_stats = overdue_result.one()

    return {
        "totalInvoices": stats.total or 0,
        "totalAmount": round(stats.total_amount or 0, 2),
        "statusDistribution": status_distribution,
        "overdueInvoices": {
            "count": overdue_stats.count or 0,
            "amount": round(overdue_stats.amount or 0, 2),
        },
        "generatedAt": datetime.now(timezone.utc).isoformat(),
    }


# =============================================================================
# GET SINGLE
# =============================================================================


@router.get(
    "/{invoice_id}",
    response_model=InvoiceTrackingResponse,
    summary="Rechnung abrufen",
    description="Ruft eine einzelne Rechnungsverfolgung ab"
)
async def get_invoice(
    invoice_id: UUID,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
    company_id: UUID = Depends(get_user_company_id_dep),
) -> InvoiceTrackingResponse:
    """Ruft eine Rechnungsverfolgung anhand der ID ab."""
    # SECURITY: Multi-Tenant RLS
    result = await db.execute(
        select(InvoiceTracking)
        .join(Document, InvoiceTracking.document_id == Document.id)
        .where(
            InvoiceTracking.id == invoice_id,
            InvoiceTracking.deleted_at.is_(None),
            Document.company_id == company_id,
        )
    )
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Rechnungsverfolgung nicht gefunden"
        )

    resp = InvoiceTrackingResponse.model_validate(invoice)

    # Compute overdue status
    now = datetime.now(timezone.utc)
    if invoice.status not in ("paid", "cancelled") and invoice.due_date:
        due_date = invoice.due_date if invoice.due_date.tzinfo else invoice.due_date.replace(tzinfo=timezone.utc)
        if now > due_date:
            resp.is_overdue = True
            resp.days_overdue = (now - due_date).days

    return resp


# =============================================================================
# CREATE
# =============================================================================


@router.post(
    "",
    response_model=InvoiceTrackingResponse,
    status_code=status.HTTP_201_CREATED,
    summary="Rechnungsverfolgung erstellen",
    description="Erstellt eine neue Rechnungsverfolgung für ein Dokument"
)
async def create_invoice(
    invoice_data: InvoiceTrackingCreate,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
    company_id: UUID = Depends(get_user_company_id_dep),
) -> InvoiceTrackingResponse:
    """
    Erstellt eine neue Rechnungsverfolgung.

    Das verknüpfte Dokument muss existieren.
    """
    # SECURITY: Prüfen ob Dokument existiert UND dem User gehoert
    doc_result = await db.execute(
        select(Document).where(
            Document.id == invoice_data.document_id,
            Document.deleted_at.is_(None),
            Document.company_id == company_id,  # Multi-Tenant RLS
        )
    )
    document = doc_result.scalar_one_or_none()

    if not document:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Verknüpftes Dokument nicht gefunden"
        )

    # Prüfen ob bereits InvoiceTracking existiert
    existing_result = await db.execute(
        select(InvoiceTracking).where(
            InvoiceTracking.document_id == invoice_data.document_id,
            InvoiceTracking.deleted_at.is_(None)
        )
    )
    if existing_result.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Für dieses Dokument existiert bereits eine Rechnungsverfolgung"
        )

    # InvoiceTracking erstellen
    invoice = InvoiceTracking(
        document_id=invoice_data.document_id,
        invoice_number=invoice_data.invoice_number,
        invoice_date=invoice_data.invoice_date,
        due_date=invoice_data.due_date,
        amount=invoice_data.amount,
        currency=invoice_data.currency,
        status=invoice_data.status.value if invoice_data.status else InvoiceStatus.OPEN.value,
    )

    db.add(invoice)
    await db.commit()
    await db.refresh(invoice)

    logger.info(
        "invoice_tracking_created",
        invoice_id=str(invoice.id),
        document_id=str(invoice.document_id),
        invoice_number=invoice.invoice_number,
    )

    return InvoiceTrackingResponse.model_validate(invoice)


# =============================================================================
# UPDATE
# =============================================================================


@router.patch(
    "/{invoice_id}",
    response_model=InvoiceTrackingResponse,
    summary="Rechnungsverfolgung aktualisieren",
    description="Aktualisiert eine Rechnungsverfolgung"
)
async def update_invoice(
    invoice_id: UUID,
    invoice_data: InvoiceTrackingUpdate,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
    company_id: UUID = Depends(get_user_company_id_dep),
) -> InvoiceTrackingResponse:
    """
    Aktualisiert eine Rechnungsverfolgung.

    Triggert automatisch Neuberechnung des Risk Scores
    wenn Status oder Zahlungsdaten geändert werden.
    """
    # SECURITY: Multi-Tenant RLS
    result = await db.execute(
        select(InvoiceTracking)
        .join(Document, InvoiceTracking.document_id == Document.id)
        .where(
            InvoiceTracking.id == invoice_id,
            InvoiceTracking.deleted_at.is_(None),
            Document.company_id == company_id,
        )
    )
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Rechnungsverfolgung nicht gefunden"
        )

    # Track if risk-relevant fields changed
    risk_relevant_changed = False
    old_status = invoice.status
    update_data = invoice_data.model_dump(exclude_unset=True)

    for field, value in update_data.items():
        if field in ("status", "paid_at", "paid_amount", "dunning_level"):
            risk_relevant_changed = True
        if field == "status" and value:
            value = value.value if hasattr(value, "value") else value
        setattr(invoice, field, value)

    invoice.updated_at = datetime.now(timezone.utc)

    await db.commit()
    await db.refresh(invoice)

    # Trigger Risk Score Recalculation
    if risk_relevant_changed:
        on_invoice_updated_recalculate.delay(str(invoice.document_id))
        logger.info(
            "invoice_update_triggered_risk_recalc",
            invoice_id=str(invoice.id),
            document_id=str(invoice.document_id),
        )

    logger.info(
        "invoice_tracking_updated",
        invoice_id=str(invoice.id),
        updated_fields=list(update_data.keys()),
    )

    # Domain Event: invoice_status_changed (wenn Status sich geaendert hat)
    if "status" in update_data and invoice.status != old_status:
        from app.services.event_sourcing.event_emitter import emit_domain_event
        await emit_domain_event(
            db=db,
            aggregate_type="invoice",
            aggregate_id=invoice.id,
            event_type="invoice_status_changed",
            event_data={
                "old_status": old_status,
                "new_status": invoice.status,
                "updated_fields": list(update_data.keys()),
            },
            company_id=company_id,
            user_id=current_user.id,
        )

    return InvoiceTrackingResponse.model_validate(invoice)


# =============================================================================
# PAYMENT STATUS UPDATE (Convenience Endpoint)
# =============================================================================


@router.post(
    "/{invoice_id}/mark-paid",
    response_model=InvoiceTrackingResponse,
    summary="Rechnung als bezahlt markieren",
    description="Markiert eine Rechnung als bezahlt und triggert Risk Score Neuberechnung"
)
async def mark_invoice_paid(
    invoice_id: UUID,
    paid_amount: Optional[float] = Query(None, description="Gezahlter Betrag (optional)"),
    paid_at: Optional[datetime] = Query(None, description="Zahlungsdatum (optional, default: jetzt)"),
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
    company_id: UUID = Depends(get_user_company_id_dep),
) -> InvoiceTrackingResponse:
    """
    Markiert eine Rechnung als bezahlt.

    - Setzt Status auf 'paid'
    - Setzt paid_at auf aktuellen Zeitpunkt (oder übergebenen Wert)
    - Setzt paid_amount auf Rechnungsbetrag (oder übergebenen Wert)
    - Triggert Risk Score Neuberechnung
    """
    # SECURITY: Multi-Tenant RLS
    result = await db.execute(
        select(InvoiceTracking)
        .join(Document, InvoiceTracking.document_id == Document.id)
        .where(
            InvoiceTracking.id == invoice_id,
            InvoiceTracking.deleted_at.is_(None),
            Document.company_id == company_id,
        )
    )
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Rechnungsverfolgung nicht gefunden"
        )

    if invoice.status == InvoiceStatus.PAID.value:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Rechnung ist bereits als bezahlt markiert"
        )

    old_status = invoice.status
    invoice.status = InvoiceStatus.PAID.value
    invoice.paid_at = paid_at or datetime.now(timezone.utc)
    invoice.paid_amount = paid_amount if paid_amount is not None else invoice.amount
    invoice.updated_at = datetime.now(timezone.utc)

    await db.commit()
    await db.refresh(invoice)

    # Trigger Risk Score Recalculation
    on_invoice_updated_recalculate.delay(str(invoice.document_id))

    logger.info(
        "invoice_marked_paid",
        invoice_id=str(invoice.id),
        document_id=str(invoice.document_id),
        paid_amount=invoice.paid_amount,
    )

    # Domain Event: invoice_status_changed (mark-paid)
    from app.services.event_sourcing.event_emitter import emit_domain_event
    await emit_domain_event(
        db=db,
        aggregate_type="invoice",
        aggregate_id=invoice.id,
        event_type="invoice_status_changed",
        event_data={
            "old_status": old_status,
            "new_status": InvoiceStatus.PAID.value,
            "paid_amount": str(invoice.paid_amount) if invoice.paid_amount else None,
        },
        company_id=company_id,
        user_id=current_user.id,
    )

    return InvoiceTrackingResponse.model_validate(invoice)


# =============================================================================
# DUNNING LEVEL UPDATE
# =============================================================================


@router.post(
    "/{invoice_id}/increase-dunning",
    response_model=InvoiceTrackingResponse,
    summary="Mahnstufe erhöhen",
    description="Erhöht die Mahnstufe einer Rechnung"
)
async def increase_dunning_level(
    invoice_id: UUID,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
    company_id: UUID = Depends(get_user_company_id_dep),
) -> InvoiceTrackingResponse:
    """
    Erhöht die Mahnstufe einer Rechnung.

    - Mahnstufe 0 -> 1: Zahlungserinnerung
    - Mahnstufe 1 -> 2: 1. Mahnung
    - Mahnstufe 2 -> 3: 2. Mahnung
    - Mahnstufe 3 -> 4: Letzte Mahnung
    - Mahnstufe 4: Maximum erreicht

    Setzt Status auf 'dunning' wenn nicht bereits.
    Triggert Risk Score Neuberechnung.
    """
    # SECURITY: Multi-Tenant RLS
    result = await db.execute(
        select(InvoiceTracking)
        .join(Document, InvoiceTracking.document_id == Document.id)
        .where(
            InvoiceTracking.id == invoice_id,
            InvoiceTracking.deleted_at.is_(None),
            Document.company_id == company_id,
        )
    )
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Rechnungsverfolgung nicht gefunden"
        )

    if invoice.status in (InvoiceStatus.PAID.value, InvoiceStatus.CANCELLED.value):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Mahnstufe kann nicht erhöht werden: Rechnung ist {invoice.status}"
        )

    if invoice.dunning_level >= 4:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Maximale Mahnstufe (4) bereits erreicht"
        )

    invoice.dunning_level += 1
    invoice.last_dunning_at = datetime.now(timezone.utc)
    invoice.status = InvoiceStatus.DUNNING.value
    invoice.updated_at = datetime.now(timezone.utc)

    await db.commit()
    await db.refresh(invoice)

    # Trigger Risk Score Recalculation
    on_invoice_updated_recalculate.delay(str(invoice.document_id))

    logger.info(
        "invoice_dunning_increased",
        invoice_id=str(invoice.id),
        document_id=str(invoice.document_id),
        new_dunning_level=invoice.dunning_level,
    )

    return InvoiceTrackingResponse.model_validate(invoice)


# =============================================================================
# DELETE (Soft Delete)
# =============================================================================


@router.delete(
    "/{invoice_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Rechnungsverfolgung löschen",
    description="Löscht eine Rechnungsverfolgung (Soft Delete)"
)
async def delete_invoice(
    invoice_id: UUID,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
    company_id: UUID = Depends(get_user_company_id_dep),
) -> None:
    """
    Löscht eine Rechnungsverfolgung (Soft Delete).

    Die Daten werden nicht physisch gelöscht, sondern nur als gelöscht markiert.
    Triggert Risk Score Neuberechnung.
    """
    # SECURITY: Multi-Tenant RLS
    result = await db.execute(
        select(InvoiceTracking)
        .join(Document, InvoiceTracking.document_id == Document.id)
        .where(
            InvoiceTracking.id == invoice_id,
            InvoiceTracking.deleted_at.is_(None),
            Document.company_id == company_id,
        )
    )
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Rechnungsverfolgung nicht gefunden"
        )

    document_id = str(invoice.document_id)

    invoice.deleted_at = datetime.now(timezone.utc)
    invoice.updated_at = datetime.now(timezone.utc)

    await db.commit()

    # Trigger Risk Score Recalculation
    on_invoice_updated_recalculate.delay(document_id)

    logger.info(
        "invoice_tracking_deleted",
        invoice_id=str(invoice_id),
        document_id=document_id,
    )


# =============================================================================
# SKONTO ENDPOINTS
# =============================================================================


@router.get(
    "/{invoice_id}/skonto",
    summary="Skonto-Informationen abrufen",
    description="Liefert Skonto-Details und -Berechnung für eine Rechnung"
)
async def get_invoice_skonto(
    invoice_id: UUID,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
    company_id: UUID = Depends(get_user_company_id_dep),
) -> dict:
    """
    Liefert Skonto-Informationen für eine Rechnung.

    **Response:**
    - skonto_percentage: Skonto-Prozentsatz
    - skonto_amount: Berechneter Skonto-Betrag
    - skonto_deadline: Ablaufdatum für Skonto
    - skonto_used: Ob Skonto bereits angewendet wurde
    - days_remaining: Verbleibende Tage
    - is_expired: Ob Skonto-Frist abgelaufen ist
    """
    from app.services.banking.skonto_service import SkontoService
    from decimal import Decimal

    # SECURITY: Multi-Tenant RLS
    result = await db.execute(
        select(InvoiceTracking)
        .join(Document, InvoiceTracking.document_id == Document.id)
        .where(
            InvoiceTracking.id == invoice_id,
            InvoiceTracking.deleted_at.is_(None),
            Document.company_id == company_id,
        )
    )
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Rechnungsverfolgung nicht gefunden"
        )

    skonto_service = SkontoService()

    # Falls Skonto-Felder vorhanden, berechnen
    if invoice.skonto_percentage and invoice.invoice_date:
        calc = await skonto_service.calculate_skonto(
            db=db,
            invoice_amount=Decimal(str(invoice.amount)),
            invoice_date=invoice.invoice_date,
            skonto_percentage=invoice.skonto_percentage,
            skonto_days=invoice.skonto_days,
            net_days=invoice.net_days,
        )
        return {
            "invoice_id": str(invoice_id),
            "skonto_percentage": calc.skonto_percentage,
            "skonto_amount": float(calc.skonto_amount),
            "skonto_deadline": calc.skonto_deadline.isoformat() if calc.skonto_deadline else None,
            "amount_with_skonto": float(calc.amount_with_skonto),
            "days_remaining": calc.days_remaining,
            "is_expired": calc.is_expired,
            "skonto_used": invoice.skonto_used,
            "savings_potential": float(calc.savings_potential),
        }
    else:
        return {
            "invoice_id": str(invoice_id),
            "skonto_percentage": None,
            "skonto_amount": None,
            "skonto_deadline": None,
            "message": "Keine Skonto-Konditionen hinterlegt",
        }


@router.patch(
    "/{invoice_id}/skonto",
    response_model=InvoiceTrackingResponse,
    summary="Skonto-Konditionen setzen",
    description="Setzt oder aktualisiert Skonto-Konditionen für eine Rechnung"
)
async def set_invoice_skonto(
    invoice_id: UUID,
    skonto_percentage: float = Query(..., ge=0, le=10, description="Skonto-Prozentsatz (0-10%)"),
    skonto_days: int = Query(10, ge=1, le=60, description="Tage für Skonto-Berechtigung"),
    net_days: int = Query(30, ge=1, le=120, description="Zahlungsziel netto"),
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
    company_id: UUID = Depends(get_user_company_id_dep),
) -> InvoiceTrackingResponse:
    """
    Setzt Skonto-Konditionen für eine Rechnung.

    Berechnet automatisch:
    - skonto_deadline
    - skonto_amount
    - due_date (aus net_days)
    """
    from app.services.banking.skonto_service import SkontoService

    # SECURITY: Multi-Tenant RLS
    result = await db.execute(
        select(InvoiceTracking)
        .join(Document, InvoiceTracking.document_id == Document.id)
        .where(
            InvoiceTracking.id == invoice_id,
            InvoiceTracking.deleted_at.is_(None),
            Document.company_id == company_id,
        )
    )
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Rechnungsverfolgung nicht gefunden"
        )

    skonto_service = SkontoService()
    # SECURITY: company_id für Defense-in-Depth Multi-Tenant Isolation
    success = await skonto_service.update_invoice_skonto_fields(
        db=db,
        invoice_tracking_id=invoice_id,
        company_id=company_id,
        skonto_percentage=skonto_percentage,
        skonto_days=skonto_days,
        net_days=net_days,
    )

    if not success:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Fehler beim Aktualisieren der Skonto-Konditionen"
        )

    await db.commit()
    await db.refresh(invoice)

    logger.info(
        "invoice_skonto_updated",
        invoice_id=str(invoice_id),
        skonto_percentage=skonto_percentage,
        skonto_days=skonto_days,
    )

    return InvoiceTrackingResponse.model_validate(invoice)


@router.post(
    "/{invoice_id}/apply-skonto",
    response_model=InvoiceTrackingResponse,
    summary="Skonto bei Zahlung anwenden",
    description="Wendet Skonto auf eine Zahlung an"
)
async def apply_invoice_skonto(
    invoice_id: UUID,
    payment_amount: float = Query(..., description="Gezahlter Betrag (mit Skonto-Abzug)"),
    payment_date: Optional[datetime] = Query(None, description="Zahlungsdatum"),
    force_apply: bool = Query(False, description="Skonto auch nach Fristablauf anwenden"),
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
    company_id: UUID = Depends(get_user_company_id_dep),
) -> InvoiceTrackingResponse:
    """
    Wendet Skonto bei einer Zahlung an.

    Prüft automatisch:
    - Ob Skonto-Konditionen vorhanden sind
    - Ob Skonto-Frist eingehalten wurde
    - Ob der Betrag zum Skonto-Abzug passt

    Mit force_apply=true kann Skonto auch nach Fristablauf angewendet werden.
    """
    from app.services.banking.skonto_service import SkontoService
    from decimal import Decimal

    # SECURITY: Multi-Tenant RLS
    result = await db.execute(
        select(InvoiceTracking)
        .join(Document, InvoiceTracking.document_id == Document.id)
        .where(
            InvoiceTracking.id == invoice_id,
            InvoiceTracking.deleted_at.is_(None),
            Document.company_id == company_id,
        )
    )
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Rechnungsverfolgung nicht gefunden"
        )

    skonto_service = SkontoService()
    # SECURITY: company_id für Defense-in-Depth Multi-Tenant Isolation
    applied, skonto_amount, message = await skonto_service.apply_skonto(
        db=db,
        invoice_tracking_id=invoice_id,
        payment_amount=Decimal(str(payment_amount)),
        payment_date=payment_date or datetime.now(timezone.utc),
        user_id=current_user.id,
        company_id=company_id,
        force_apply=force_apply,
    )

    await db.commit()
    await db.refresh(invoice)

    if not applied:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=message
        )

    # Trigger Risk Score Recalculation
    on_invoice_updated_recalculate.delay(str(invoice.document_id))

    logger.info(
        "invoice_skonto_applied",
        invoice_id=str(invoice_id),
        skonto_amount=float(skonto_amount),
        message=message,
    )

    return InvoiceTrackingResponse.model_validate(invoice)


@router.get(
    "/skonto/upcoming",
    summary="Anstehende Skonto-Fristen",
    description="Listet Rechnungen mit bald ablaufenden Skonto-Fristen"
)
async def get_upcoming_skonto_deadlines(
    days_ahead: int = Query(7, ge=1, le=30, description="Tage im Voraus"),
    limit: int = Query(20, ge=1, le=100, description="Maximale Anzahl"),
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
    company_id: UUID = Depends(get_user_company_id_dep),
) -> list:
    """
    Listet Rechnungen mit anstehenden Skonto-Fristen.

    **Response:**
    - Sortiert nach Dringlichkeit (kritisch < 1 Tag, warning < 3 Tage)
    - Zeigt potenzielle Ersparnis
    """
    from app.services.banking.skonto_service import SkontoService

    skonto_service = SkontoService()
    alerts = await skonto_service.get_upcoming_skonto_deadlines(
        db=db,
        company_id=company_id,
        days_ahead=days_ahead,
        limit=limit,
    )

    return [
        {
            "invoice_id": str(alert.invoice_id),
            "invoice_number": alert.invoice_number,
            "entity_name": alert.entity_name,
            "skonto_deadline": alert.skonto_deadline.isoformat(),
            "skonto_amount": float(alert.skonto_amount),
            "days_remaining": alert.days_remaining,
            "urgency": alert.urgency,
        }
        for alert in alerts
    ]


@router.get(
    "/skonto/missed",
    summary="Verpasste Skonto-Möglichkeiten",
    description="Listet alle Rechnungen mit verpassten Skonto-Fristen"
)
async def get_missed_skonto(
    start_date: Optional[datetime] = Query(None, description="Startdatum (ISO 8601)"),
    end_date: Optional[datetime] = Query(None, description="Enddatum (ISO 8601)"),
    page: int = Query(1, ge=1, description="Seitennummer"),
    per_page: int = Query(20, ge=1, le=100, description="Einträge pro Seite"),
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
    company_id: UUID = Depends(get_user_company_id_dep),
) -> dict:
    """
    Listet alle Rechnungen mit verpassten Skonto-Möglichkeiten.

    **Response:**
    - items: Liste der verpassten Skonto-Rechnungen
    - total: Gesamtanzahl
    - total_missed_amount: Summe verpasster Ersparnisse
    """
    from app.db.models import BusinessEntity
    from app.core.datetime_utils import utc_now
    from decimal import Decimal

    now = utc_now()

    # Basis-Filter: Verpasste Skonto-Rechnungen
    base_conditions = [
        InvoiceTracking.company_id == company_id,
        InvoiceTracking.skonto_percentage.isnot(None),
        InvoiceTracking.skonto_percentage > 0,
        InvoiceTracking.skonto_deadline.isnot(None),
        InvoiceTracking.skonto_deadline < now,
        InvoiceTracking.skonto_used == False,
        InvoiceTracking.deleted_at.is_(None),
    ]

    # Optionale Datumsfilter
    if start_date:
        base_conditions.append(InvoiceTracking.invoice_date >= start_date)
    if end_date:
        base_conditions.append(InvoiceTracking.invoice_date <= end_date)

    # Count Query
    count_stmt = select(func.count(InvoiceTracking.id)).where(and_(*base_conditions))
    count_result = await db.execute(count_stmt)
    total = count_result.scalar() or 0

    # Summe der verpassten Betraege
    sum_stmt = select(
        func.sum(
            InvoiceTracking.amount * InvoiceTracking.skonto_percentage / 100
        )
    ).where(and_(*base_conditions))
    sum_result = await db.execute(sum_stmt)
    total_missed_amount = sum_result.scalar() or 0.0

    # Daten-Query mit Pagination
    offset = (page - 1) * per_page
    data_stmt = (
        select(InvoiceTracking, Document, BusinessEntity)
        .join(Document, InvoiceTracking.document_id == Document.id)
        .outerjoin(BusinessEntity, Document.business_entity_id == BusinessEntity.id)
        .where(and_(*base_conditions))
        .order_by(InvoiceTracking.skonto_deadline.desc())
        .offset(offset)
        .limit(per_page)
    )
    data_result = await db.execute(data_stmt)
    rows = data_result.all()

    items = []
    for invoice, document, entity in rows:
        skonto_amount = float(
            Decimal(str(invoice.amount)) * Decimal(str(invoice.skonto_percentage)) / Decimal("100")
        )
        days_missed = (now - invoice.skonto_deadline).days if invoice.skonto_deadline else 0

        items.append({
            "invoice_id": str(invoice.id),
            "invoice_number": invoice.invoice_number or document.original_filename,
            "document_id": str(invoice.document_id),
            "entity_id": str(entity.id) if entity else None,
            "entity_name": entity.name if entity else "Unbekannt",
            "invoice_date": invoice.invoice_date.isoformat() if invoice.invoice_date else None,
            "amount": float(invoice.amount),
            "skonto_percentage": invoice.skonto_percentage,
            "skonto_amount": round(skonto_amount, 2),
            "skonto_deadline": invoice.skonto_deadline.isoformat() if invoice.skonto_deadline else None,
            "days_missed_by": days_missed,
            "paid_at": invoice.paid_at.isoformat() if invoice.paid_at else None,
            "paid_amount": float(invoice.paid_amount) if invoice.paid_amount else None,
        })

    return {
        "items": items,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_missed_amount": round(float(total_missed_amount), 2),
    }


@router.get(
    "/skonto/statistics",
    summary="Skonto-Statistiken",
    description="Berechnet Skonto-Statistiken für einen Zeitraum"
)
async def get_skonto_statistics(
    start_date: datetime = Query(..., description="Startdatum (ISO 8601)"),
    end_date: datetime = Query(..., description="Enddatum (ISO 8601)"),
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
    company_id: UUID = Depends(get_user_company_id_dep),
) -> dict:
    """
    Berechnet Skonto-Statistiken für einen Zeitraum.

    **Response:**
    - period_start/period_end: Zeitraum
    - total_invoices: Gesamtzahl Rechnungen
    - invoices_with_skonto: Rechnungen mit Skonto
    - skonto_used_count: Skonto genutzt
    - skonto_missed_count: Skonto verpasst
    - skonto_pending_count: Skonto noch offen
    - total_savings: Gesparte Betraege
    - missed_savings: Verpasste Ersparnisse
    - potential_savings: Potentielle Ersparnisse
    - usage_rate: Nutzungsrate in Prozent
    """
    from app.services.banking.skonto_service import SkontoService

    skonto_service = SkontoService()
    stats = await skonto_service.get_skonto_statistics(
        db=db,
        company_id=company_id,
        start_date=start_date,
        end_date=end_date,
    )

    return {
        "period_start": stats.period_start.isoformat(),
        "period_end": stats.period_end.isoformat(),
        "total_invoices": stats.total_invoices,
        "invoices_with_skonto": stats.invoices_with_skonto,
        "skonto_used_count": stats.skonto_used_count,
        "skonto_missed_count": stats.skonto_missed_count,
        "skonto_pending_count": stats.skonto_pending_count,
        "total_savings": float(stats.total_savings),
        "missed_savings": float(stats.missed_savings),
        "potential_savings": float(stats.potential_savings),
        "usage_rate": stats.usage_rate,
    }


@router.get(
    "/skonto/monthly-summary",
    summary="Monatliche Skonto-Übersicht",
    description="Liefert monatliche Skonto-Zusammenfassung für Chart-Darstellung"
)
async def get_monthly_skonto_summary(
    months: int = Query(12, ge=1, le=24, description="Anzahl Monate"),
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
    company_id: UUID = Depends(get_user_company_id_dep),
) -> list:
    """
    Liefert monatliche Skonto-Zusammenfassung.

    **Response:**
    - List von Monaten mit:
      - year, month: Jahr und Monat
      - used_amount: Genutzte Skonto-Betraege
      - missed_amount: Verpasste Skonto-Betraege
      - usage_rate: Nutzungsrate
    """
    from app.core.datetime_utils import utc_now
    from datetime import timedelta
    from decimal import Decimal
    from calendar import monthrange

    now = utc_now()
    results = []

    # Für jeden Monat Statistiken berechnen
    for i in range(months - 1, -1, -1):
        # Monat berechnen (rückwärts)
        target_date = now - timedelta(days=i * 30)
        year = target_date.year
        month = target_date.month

        # Monatsanfang und -ende
        _, last_day = monthrange(year, month)
        month_start = datetime(year, month, 1, tzinfo=timezone.utc)
        month_end = datetime(year, month, last_day, 23, 59, 59, tzinfo=timezone.utc)

        # Query für diesen Monat
        month_conditions = [
            InvoiceTracking.company_id == company_id,
            InvoiceTracking.skonto_percentage.isnot(None),
            InvoiceTracking.skonto_percentage > 0,
            InvoiceTracking.invoice_date >= month_start,
            InvoiceTracking.invoice_date <= month_end,
            InvoiceTracking.deleted_at.is_(None),
        ]

        # Alle Rechnungen des Monats mit Skonto
        stmt = select(InvoiceTracking).where(and_(*month_conditions))
        result = await db.execute(stmt)
        invoices = result.scalars().all()

        used_amount = Decimal("0.00")
        missed_amount = Decimal("0.00")
        used_count = 0
        missed_count = 0

        for inv in invoices:
            skonto_amount = Decimal(str(inv.amount)) * Decimal(str(inv.skonto_percentage)) / Decimal("100")
            skonto_amount = skonto_amount.quantize(Decimal("0.01"))

            if inv.skonto_used:
                used_amount += skonto_amount
                used_count += 1
            elif inv.skonto_deadline and inv.skonto_deadline < now:
                missed_amount += skonto_amount
                missed_count += 1

        total_with_outcome = used_count + missed_count
        usage_rate = (used_count / total_with_outcome * 100) if total_with_outcome > 0 else 0.0

        results.append({
            "year": str(year),
            "month": str(month).zfill(2),
            "used_amount": float(used_amount),
            "missed_amount": float(missed_amount),
            "used_count": used_count,
            "missed_count": missed_count,
            "usage_rate": round(usage_rate, 1),
        })

    return results


@router.get(
    "/skonto/missed/export",
    summary="Verpasste Skonto exportieren",
    description="Exportiert verpasste Skonto-Daten als Excel oder CSV"
)
async def export_missed_skonto(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$", description="Export-Format"),
    start_date: Optional[datetime] = Query(None, description="Startdatum"),
    end_date: Optional[datetime] = Query(None, description="Enddatum"),
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
    company_id: UUID = Depends(get_user_company_id_dep),
):
    """
    Exportiert verpasste Skonto-Daten.

    **Formate:**
    - xlsx: Excel-Datei
    - csv: CSV-Datei
    """
    from fastapi.responses import StreamingResponse
    from app.db.models import BusinessEntity
    from app.core.datetime_utils import utc_now
    from decimal import Decimal
    import io

    now = utc_now()

    # Filter
    conditions = [
        InvoiceTracking.company_id == company_id,
        InvoiceTracking.skonto_percentage.isnot(None),
        InvoiceTracking.skonto_percentage > 0,
        InvoiceTracking.skonto_deadline.isnot(None),
        InvoiceTracking.skonto_deadline < now,
        InvoiceTracking.skonto_used == False,
        InvoiceTracking.deleted_at.is_(None),
    ]

    if start_date:
        conditions.append(InvoiceTracking.invoice_date >= start_date)
    if end_date:
        conditions.append(InvoiceTracking.invoice_date <= end_date)

    stmt = (
        select(InvoiceTracking, Document, BusinessEntity)
        .join(Document, InvoiceTracking.document_id == Document.id)
        .outerjoin(BusinessEntity, Document.business_entity_id == BusinessEntity.id)
        .where(and_(*conditions))
        .order_by(InvoiceTracking.skonto_deadline.desc())
    )
    result = await db.execute(stmt)
    rows = result.all()

    # Daten aufbereiten
    data = []
    for invoice, document, entity in rows:
        skonto_amount = float(
            Decimal(str(invoice.amount)) * Decimal(str(invoice.skonto_percentage)) / Decimal("100")
        )
        days_missed = (now - invoice.skonto_deadline).days if invoice.skonto_deadline else 0

        data.append({
            "Rechnungsnummer": invoice.invoice_number or document.original_filename,
            "Geschäftspartner": entity.name if entity else "Unbekannt",
            "Rechnungsdatum": invoice.invoice_date.strftime("%d.%m.%Y") if invoice.invoice_date else "",
            "Rechnungsbetrag": f"{invoice.amount:.2f}".replace(".", ","),
            "Skonto %": f"{invoice.skonto_percentage:.1f}".replace(".", ","),
            "Skonto Betrag": f"{skonto_amount:.2f}".replace(".", ","),
            "Skonto Frist": invoice.skonto_deadline.strftime("%d.%m.%Y") if invoice.skonto_deadline else "",
            "Tage verpasst": days_missed,
            "Bezahlt am": invoice.paid_at.strftime("%d.%m.%Y") if invoice.paid_at else "",
        })

    if format == "csv":
        import csv
        output = io.StringIO()
        if data:
            writer = csv.DictWriter(output, fieldnames=data[0].keys(), delimiter=";")
            writer.writeheader()
            writer.writerows(data)
        content = output.getvalue().encode("utf-8-sig")  # BOM für Excel
        media_type = "text/csv; charset=utf-8"
        filename = f"verpasste_skonto_{now.strftime('%Y%m%d')}.csv"
    else:
        # Excel mit openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Verpasste Skonto"

            # Header
            headers = list(data[0].keys()) if data else []
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Daten
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, key in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data[key])

            # Spaltenbreiten anpassen
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception as e:
                        logger.debug(
                            "excel_cell_length_calculation_failed",
                            error_type=type(e).__name__,
                        )
                ws.column_dimensions[column].width = min(max_length + 2, 50)

            output = io.BytesIO()
            wb.save(output)
            content = output.getvalue()
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"verpasste_skonto_{now.strftime('%Y%m%d')}.xlsx"
        except ImportError:
            # Fallback zu CSV wenn openpyxl nicht installiert
            import csv
            output = io.StringIO()
            if data:
                writer = csv.DictWriter(output, fieldnames=data[0].keys(), delimiter=";")
                writer.writeheader()
                writer.writerows(data)
            content = output.getvalue().encode("utf-8-sig")
            media_type = "text/csv; charset=utf-8"
            filename = f"verpasste_skonto_{now.strftime('%Y%m%d')}.csv"

    logger.info(
        "missed_skonto_export",
        format=format,
        records=len(data),
        user_id=str(current_user.id),
    )

    return StreamingResponse(
        io.BytesIO(content) if isinstance(content, bytes) else io.BytesIO(content.encode()),
        media_type=media_type,
        headers={"Content-Disposition": build_content_disposition(filename, "attachment")},
    )


# =============================================================================
# PARTIAL PAYMENT ENDPOINTS
# =============================================================================


@router.post(
    "/{invoice_id}/payments",
    status_code=status.HTTP_201_CREATED,
    summary="Teilzahlung erfassen",
    description="Erfasst eine Teilzahlung für eine Rechnung"
)
async def record_partial_payment(
    invoice_id: UUID,
    amount: float = Query(..., gt=0, description="Zahlungsbetrag"),
    payment_reference: Optional[str] = Query(None, description="Verwendungszweck/Referenz"),
    payment_method: str = Query("bank_transfer", description="Zahlungsmethode"),
    transaction_date: Optional[datetime] = Query(None, description="Zahlungsdatum"),
    notes: Optional[str] = Query(None, description="Interne Notizen"),
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
    company_id: UUID = Depends(get_user_company_id_dep),
) -> dict:
    """
    Erfasst eine Teilzahlung.

    **Zahlungsmethoden:**
    - bank_transfer (Standard)
    - credit_card
    - cash
    - sepa_direct_debit
    - paypal

    Aktualisiert automatisch:
    - paid_amount (Summe aller Zahlungen)
    - outstanding_amount (Restbetrag)
    - Status (partial oder paid)
    """
    from app.services.banking.partial_payment_service import (
        PartialPaymentService,
        PaymentTransactionCreate,
    )
    from decimal import Decimal

    # SECURITY: Multi-Tenant RLS
    result = await db.execute(
        select(InvoiceTracking)
        .join(Document, InvoiceTracking.document_id == Document.id)
        .where(
            InvoiceTracking.id == invoice_id,
            InvoiceTracking.deleted_at.is_(None),
            Document.company_id == company_id,
        )
    )
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Rechnungsverfolgung nicht gefunden"
        )

    payment_service = PartialPaymentService()
    payment_data = PaymentTransactionCreate(
        amount=Decimal(str(amount)),
        transaction_date=transaction_date,
        payment_reference=payment_reference,
        payment_method=payment_method,
        notes=notes,
    )

    try:
        transaction, message = await payment_service.record_payment(
            db=db,
            invoice_tracking_id=invoice_id,
            payment_data=payment_data,
            user_id=current_user.id,
            company_id=company_id,
        )
    except ValueError as e:
        logger.error("payment_recording_failed", **safe_error_log(e, context="Teilzahlung erfassen"))
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=safe_error_detail(e, "Teilzahlung erfassen")
        )

    await db.commit()

    # Trigger Risk Score Recalculation
    on_invoice_updated_recalculate.delay(str(invoice.document_id))

    logger.info(
        "partial_payment_recorded",
        invoice_id=str(invoice_id),
        payment_id=str(transaction.id),
        amount=float(amount),
    )

    return {
        "payment_id": str(transaction.id),
        "invoice_id": str(invoice_id),
        "amount": float(transaction.amount),
        "transaction_date": transaction.transaction_date.isoformat(),
        "payment_method": transaction.payment_method,
        "reconciliation_status": transaction.reconciliation_status,
        "message": message,
    }


@router.get(
    "/{invoice_id}/payments",
    summary="Zahlungen einer Rechnung abrufen",
    description="Listet alle Zahlungen für eine Rechnung"
)
async def get_invoice_payments(
    invoice_id: UUID,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
    company_id: UUID = Depends(get_user_company_id_dep),
) -> dict:
    """
    Liefert Zahlungsübersicht für eine Rechnung.

    **Response:**
    - total_amount: Rechnungsbetrag
    - paid_amount: Summe aller Zahlungen
    - outstanding_amount: Ausstehender Betrag
    - payment_count: Anzahl Zahlungen
    - payments: Liste der einzelnen Zahlungen
    - is_fully_paid: Ob vollständig bezahlt
    """
    from app.services.banking.partial_payment_service import PartialPaymentService

    # SECURITY: Multi-Tenant RLS
    result = await db.execute(
        select(InvoiceTracking)
        .join(Document, InvoiceTracking.document_id == Document.id)
        .where(
            InvoiceTracking.id == invoice_id,
            InvoiceTracking.deleted_at.is_(None),
            Document.company_id == company_id,
        )
    )
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Rechnungsverfolgung nicht gefunden"
        )

    payment_service = PartialPaymentService()

    try:
        # SECURITY: company_id für Multi-Tenant Isolation
        summary = await payment_service.get_payment_summary(
            db=db,
            invoice_tracking_id=invoice_id,
            company_id=company_id,
        )
    except ValueError as e:
        logger.error("payment_summary_failed", **safe_error_log(e, context="Zahlungsübersicht abrufen"))
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=safe_error_detail(e, "Zahlungsübersicht abrufen")
        )

    return {
        "invoice_id": str(summary.invoice_tracking_id),
        "invoice_number": summary.invoice_number,
        "total_amount": float(summary.total_amount),
        "paid_amount": float(summary.paid_amount),
        "outstanding_amount": float(summary.outstanding_amount),
        "skonto_total": float(summary.skonto_total),
        "payment_count": summary.payment_count,
        "is_fully_paid": summary.is_fully_paid,
        "overpaid_amount": float(summary.overpaid_amount),
        "payments": [
            {
                "id": str(p.id),
                "amount": float(p.amount),
                "transaction_date": p.transaction_date.isoformat(),
                "payment_reference": p.payment_reference,
                "payment_method": p.payment_method,
                "skonto_deducted": float(p.skonto_deducted) if p.skonto_deducted else None,
                "reconciliation_status": p.reconciliation_status,
                "created_at": p.created_at.isoformat(),
            }
            for p in summary.payments
        ],
    }


@router.delete(
    "/{invoice_id}/payments/{payment_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Teilzahlung löschen",
    description="Löscht eine Teilzahlung (nur wenn nicht reconciled)"
)
async def delete_partial_payment(
    invoice_id: UUID,
    payment_id: UUID,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
    company_id: UUID = Depends(get_user_company_id_dep),
) -> None:
    """
    Löscht eine Teilzahlung.

    Nur möglich wenn:
    - Zahlung noch nicht mit Bank-Transaktion verknüpft
    - Benutzer berechtigt
    """
    from app.services.banking.partial_payment_service import PartialPaymentService

    # SECURITY: Multi-Tenant RLS - Prüfen ob Invoice dem User gehoert
    result = await db.execute(
        select(InvoiceTracking)
        .join(Document, InvoiceTracking.document_id == Document.id)
        .where(
            InvoiceTracking.id == invoice_id,
            InvoiceTracking.deleted_at.is_(None),
            Document.company_id == company_id,
        )
    )
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Rechnungsverfolgung nicht gefunden"
        )

    payment_service = PartialPaymentService()
    # SECURITY: company_id für Multi-Tenant Isolation
    success, message = await payment_service.delete_payment(
        db=db,
        payment_transaction_id=payment_id,
        user_id=current_user.id,
        company_id=company_id,
    )

    if not success:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=message
        )

    await db.commit()

    # Trigger Risk Score Recalculation
    on_invoice_updated_recalculate.delay(str(invoice.document_id))

    logger.info(
        "partial_payment_deleted",
        invoice_id=str(invoice_id),
        payment_id=str(payment_id),
    )
