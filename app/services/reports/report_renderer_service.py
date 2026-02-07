# -*- coding: utf-8 -*-
"""
Report Renderer Service.

Rendert Reports in verschiedene Formate: PDF, Excel, CSV, JSON.
"""

from __future__ import annotations

import csv
import io
import json
import uuid
from datetime import datetime, timezone
from typing import Dict, List, Optional

import structlog

from app.services.reports.report_builder_service import ReportResult

logger = structlog.get_logger(__name__)


# Versuche openpyxl zu importieren
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available - Excel export disabled")

# Versuche reportlab zu importieren
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logger.warning("reportlab not available - PDF export disabled")


class ReportRendererService:
    """Service fuer Report-Rendering in verschiedene Formate."""

    _instance: Optional["ReportRendererService"] = None

    def __new__(cls) -> "ReportRendererService":
        """Singleton-Pattern."""
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance

    def get_supported_formats(self) -> List[Dict[str, object]]:
        """Gibt unterstuetzte Export-Formate zurueck."""
        formats = [
            {"id": "json", "name": "JSON", "extension": ".json", "mime_type": "application/json", "available": True},
            {"id": "csv", "name": "CSV", "extension": ".csv", "mime_type": "text/csv", "available": True},
        ]

        if OPENPYXL_AVAILABLE:
            formats.append({
                "id": "excel",
                "name": "Excel",
                "extension": ".xlsx",
                "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "available": True,
            })
        else:
            formats.append({"id": "excel", "name": "Excel", "extension": ".xlsx", "available": False})

        if REPORTLAB_AVAILABLE:
            formats.append({
                "id": "pdf",
                "name": "PDF",
                "extension": ".pdf",
                "mime_type": "application/pdf",
                "available": True,
            })
        else:
            formats.append({"id": "pdf", "name": "PDF", "extension": ".pdf", "available": False})

        return formats

    async def render(
        self,
        result: ReportResult,
        format: str,
        layout_config: Optional[Dict[str, object]] = None,
        chart_configs: Optional[List[Dict[str, object]]] = None,
    ) -> bytes:
        """Rendert einen Report in das angegebene Format."""
        if format == "json":
            return await self.render_json(result)
        elif format == "csv":
            return await self.render_csv(result)
        elif format == "excel":
            return await self.render_excel(result, layout_config, chart_configs)
        elif format == "pdf":
            return await self.render_pdf(result, layout_config)
        else:
            raise ValueError(f"Unbekanntes Format: {format}")

    async def render_json(self, result: ReportResult) -> bytes:
        """Rendert Report als JSON."""
        output = {
            "template_id": str(result.template_id),
            "template_name": result.template_name,
            "executed_at": result.executed_at.isoformat(),
            "total_count": result.total_count,
            "columns": result.columns,
            "rows": [row.data for row in result.rows],
            "aggregations": result.aggregations,
            "filters_applied": result.filters_applied,
        }

        return json.dumps(output, ensure_ascii=False, indent=2, default=str).encode("utf-8")

    async def render_csv(self, result: ReportResult) -> bytes:
        """Rendert Report als CSV."""
        output = io.StringIO()
        writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_MINIMAL)

        # Header
        headers = [col["display_name"] for col in result.columns]
        writer.writerow(headers)

        # Daten
        for row in result.rows:
            row_values = []
            for col in result.columns:
                value = row.data.get(col["field_path"], "")
                row_values.append(self._format_csv_value(value, col.get("data_type", "string")))
            writer.writerow(row_values)

        # Aggregationen als letzte Zeile
        if result.aggregations:
            agg_row = []
            for col in result.columns:
                agg = result.aggregations.get(col["field_path"])
                if agg:
                    agg_row.append(f"{agg['type']}: {agg['value']}")
                else:
                    agg_row.append("")
            writer.writerow(agg_row)

        return output.getvalue().encode("utf-8-sig")  # BOM fuer Excel-Kompatibilitaet

    async def render_excel(
        self,
        result: ReportResult,
        layout_config: Optional[Dict[str, object]] = None,
        chart_configs: Optional[List[Dict[str, object]]] = None,
    ) -> bytes:
        """Rendert Report als Excel-Datei."""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl nicht installiert - Excel-Export nicht verfuegbar")

        wb = Workbook()
        ws = wb.active
        ws.title = result.template_name[:31]  # Excel max 31 chars

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        currency_style = "#,##0.00 €"
        date_style = "DD.MM.YYYY"

        # Titel
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(result.columns))
        title_cell = ws.cell(row=1, column=1, value=result.template_name)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        # Metadaten
        ws.cell(row=2, column=1, value=f"Erstellt: {result.executed_at.strftime('%d.%m.%Y %H:%M')}")
        ws.cell(row=2, column=2, value=f"Datensaetze: {result.total_count}")

        # Header (Zeile 4)
        header_row = 4
        for col_idx, col in enumerate(result.columns, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=col["display_name"])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Daten
        for row_idx, row in enumerate(result.rows, start=header_row + 1):
            for col_idx, col in enumerate(result.columns, start=1):
                value = row.data.get(col["field_path"])
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                # Format anwenden
                data_type = col.get("data_type", "string")
                if data_type == "currency" and value is not None:
                    cell.number_format = currency_style
                elif data_type == "date" and value is not None:
                    cell.number_format = date_style
                elif data_type == "number" and value is not None:
                    cell.number_format = "#,##0.00"

        # Aggregationszeile
        if result.aggregations:
            agg_row = header_row + len(result.rows) + 1
            for col_idx, col in enumerate(result.columns, start=1):
                agg = result.aggregations.get(col["field_path"])
                if agg:
                    cell = ws.cell(row=agg_row, column=col_idx, value=agg["value"])
                    cell.font = Font(bold=True)
                    cell.border = thin_border

                    data_type = col.get("data_type", "string")
                    if data_type == "currency":
                        cell.number_format = currency_style

        # Spaltenbreiten automatisch anpassen
        for col_idx, col in enumerate(result.columns, start=1):
            column_letter = get_column_letter(col_idx)
            # Berechne Breite basierend auf Header und max 50 Zeichen
            max_length = len(col["display_name"])
            for row in result.rows[:100]:  # Nur erste 100 Zeilen pruefen
                value = row.data.get(col["field_path"])
                if value:
                    max_length = max(max_length, min(len(str(value)), 50))
            ws.column_dimensions[column_letter].width = max_length + 2

        # Charts hinzufuegen (falls konfiguriert)
        if chart_configs:
            await self._add_excel_charts(ws, result, chart_configs, header_row)

        # In Bytes konvertieren
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    async def render_pdf(
        self,
        result: ReportResult,
        layout_config: Optional[Dict[str, object]] = None,
    ) -> bytes:
        """Rendert Report als PDF."""
        if not REPORTLAB_AVAILABLE:
            raise RuntimeError("reportlab nicht installiert - PDF-Export nicht verfuegbar")

        # Layout-Konfiguration
        layout_config = layout_config or {}
        orientation = layout_config.get("orientation", "portrait")
        page_size = landscape(A4) if orientation == "landscape" else A4
        margins = layout_config.get("margins", {"left": 2, "right": 2, "top": 2, "bottom": 2})

        # PDF erstellen
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=page_size,
            leftMargin=margins.get("left", 2) * cm,
            rightMargin=margins.get("right", 2) * cm,
            topMargin=margins.get("top", 2) * cm,
            bottomMargin=margins.get("bottom", 2) * cm,
        )

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Title",
            parent=styles["Heading1"],
            fontSize=16,
            spaceAfter=12,
        )
        meta_style = ParagraphStyle(
            "Meta",
            parent=styles["Normal"],
            fontSize=10,
            textColor=colors.grey,
            spaceAfter=12,
        )

        # Content erstellen
        elements = []

        # Titel
        elements.append(Paragraph(result.template_name, title_style))

        # Metadaten
        meta_text = f"Erstellt: {result.executed_at.strftime('%d.%m.%Y %H:%M')} | Datensaetze: {result.total_count}"
        elements.append(Paragraph(meta_text, meta_style))
        elements.append(Spacer(1, 12))

        # Tabelle erstellen
        table_data = []

        # Header
        headers = [col["display_name"] for col in result.columns]
        table_data.append(headers)

        # Daten (max 1000 Zeilen fuer PDF)
        for row in result.rows[:1000]:
            row_values = []
            for col in result.columns:
                value = row.data.get(col["field_path"], "")
                row_values.append(self._format_pdf_value(value, col.get("data_type", "string")))
            table_data.append(row_values)

        # Aggregationen
        if result.aggregations:
            agg_row = []
            for col in result.columns:
                agg = result.aggregations.get(col["field_path"])
                if agg:
                    agg_row.append(f"{agg['type']}: {self._format_pdf_value(agg['value'], col.get('data_type', 'string'))}")
                else:
                    agg_row.append("")
            table_data.append(agg_row)

        # Spaltenbreiten berechnen
        num_cols = len(result.columns)
        available_width = page_size[0] - (margins.get("left", 2) + margins.get("right", 2)) * cm
        col_widths = [available_width / num_cols] * num_cols

        # Tabelle erstellen
        table = Table(table_data, colWidths=col_widths)

        # Tabellen-Style
        style = TableStyle([
            # Header
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E86AB")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),

            # Daten
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (0, 1), (-1, -1), "LEFT"),

            # Rahmen
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),

            # Alternating row colors
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ])

        # Aggregationszeile hervorheben
        if result.aggregations:
            style.add("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8E8E8"))
            style.add("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold")

        table.setStyle(style)
        elements.append(table)

        # Hinweis wenn mehr als 1000 Zeilen
        if len(result.rows) > 1000:
            elements.append(Spacer(1, 12))
            elements.append(Paragraph(
                f"Hinweis: Es werden nur die ersten 1000 von {len(result.rows)} Zeilen angezeigt.",
                meta_style
            ))

        # PDF generieren
        doc.build(elements)
        output.seek(0)

        return output.getvalue()

    # =========================================================================
    # HELPER METHODS
    # =========================================================================

    def _format_csv_value(self, value: object, data_type: str) -> str:
        """Formatiert einen Wert fuer CSV-Export."""
        if value is None:
            return ""

        if data_type == "currency":
            try:
                return f"{float(value):.2f}".replace(".", ",")
            except (ValueError, TypeError):
                return str(value)
        elif data_type == "date":
            if isinstance(value, str) and "T" in value:
                # ISO-Format zu deutschem Format
                try:
                    dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
                    return dt.strftime("%d.%m.%Y")
                except ValueError:
                    return value
            return str(value)
        elif data_type == "number":
            try:
                return f"{float(value):.2f}".replace(".", ",")
            except (ValueError, TypeError):
                return str(value)

        return str(value)

    def _format_pdf_value(self, value: object, data_type: str) -> str:
        """Formatiert einen Wert fuer PDF-Export."""
        if value is None:
            return "-"

        if data_type == "currency":
            try:
                return f"{float(value):,.2f} EUR".replace(",", "X").replace(".", ",").replace("X", ".")
            except (ValueError, TypeError):
                return str(value)
        elif data_type == "date":
            if isinstance(value, str) and "T" in value:
                try:
                    dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
                    return dt.strftime("%d.%m.%Y")
                except ValueError:
                    return value
            return str(value)
        elif data_type == "number":
            try:
                return f"{float(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            except (ValueError, TypeError):
                return str(value)

        return str(value) if value else "-"

    async def _add_excel_charts(
        self,
        ws: object,
        result: ReportResult,
        chart_configs: List[Dict[str, object]],
        data_start_row: int,
    ) -> None:
        """Fuegt Charts zu einem Excel-Worksheet hinzu."""
        if not OPENPYXL_AVAILABLE:
            return

        data_end_row = data_start_row + len(result.rows)

        for chart_idx, chart_config in enumerate(chart_configs):
            chart_type = chart_config.get("chart_type", "bar")
            title = chart_config.get("title", f"Chart {chart_idx + 1}")
            y_axis_fields = chart_config.get("y_axis_fields", [])

            # Chart-Objekt erstellen
            if chart_type == "bar":
                chart = BarChart()
            elif chart_type == "line":
                chart = LineChart()
            elif chart_type == "pie":
                chart = PieChart()
            else:
                chart = BarChart()

            chart.title = title
            chart.style = 10

            # Daten-Referenzen erstellen
            for y_field in y_axis_fields:
                # Finde Spalten-Index
                col_idx = None
                for idx, col in enumerate(result.columns, start=1):
                    if col["field_path"] == y_field:
                        col_idx = idx
                        break

                if col_idx:
                    data = Reference(
                        ws,
                        min_col=col_idx,
                        min_row=data_start_row,
                        max_row=data_end_row,
                    )
                    chart.add_data(data, titles_from_data=True)

            # Chart positionieren
            position = chart_config.get("position", "bottom")
            if position == "bottom":
                chart_row = data_end_row + 3
            else:
                chart_row = data_start_row

            ws.add_chart(chart, f"A{chart_row}")
