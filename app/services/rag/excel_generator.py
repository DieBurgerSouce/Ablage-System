"""Excel Report Generator.

Generiert Excel-Reports mit:
- Multi-Sheet Support
- Charts und Grafiken
- Formatierung und Styling
- LLM-gestuetzte Inhalte
"""

import structlog
from typing import List, Dict, Any, Optional
from pathlib import Path
from datetime import datetime
from io import BytesIO

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
        NamedStyle
    )
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = structlog.get_logger(__name__)


class ExcelReportGenerator:
    """Generator fuer Excel-Reports."""

    # Styles
    HEADER_FONT = Font(bold=True, size=12, color="FFFFFF") if OPENPYXL_AVAILABLE else None
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") if OPENPYXL_AVAILABLE else None
    TITLE_FONT = Font(bold=True, size=16) if OPENPYXL_AVAILABLE else None
    SUBTITLE_FONT = Font(bold=True, size=14, color="4472C4") if OPENPYXL_AVAILABLE else None

    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl_not_available")

    def create_report(
        self,
        title: str,
        data: Dict[str, Any],
        output_path: Optional[Path] = None
    ) -> bytes:
        """
        Erstellt einen Excel-Report.

        Args:
            title: Report-Titel
            data: Report-Daten mit Sheets und Charts
            output_path: Optionaler Speicherpfad

        Returns:
            Excel-Datei als Bytes
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl ist nicht installiert. "
                "Installieren mit: pip install openpyxl"
            )

        wb = Workbook()

        # Standard-Sheet entfernen
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Summary Sheet erstellen
        self._create_summary_sheet(wb, title, data.get("summary", {}))

        # Daten-Sheets erstellen
        for sheet_name, sheet_data in data.get("sheets", {}).items():
            self._create_data_sheet(wb, sheet_name, sheet_data)

        # Charts hinzufuegen
        for chart_config in data.get("charts", []):
            self._add_chart(wb, chart_config)

        # Als Bytes speichern
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        content = output.read()

        # Optional auf Disk speichern
        if output_path:
            output_path.write_bytes(content)
            logger.info("excel_report_saved", path=str(output_path))

        return content

    def _create_summary_sheet(
        self,
        wb: "Workbook",
        title: str,
        summary: Dict[str, Any]
    ):
        """Erstellt das Summary-Sheet."""
        ws = wb.create_sheet("Zusammenfassung", 0)

        # Titel
        ws["A1"] = title
        ws["A1"].font = self.TITLE_FONT
        ws.merge_cells("A1:F1")

        # Generierungsdatum
        ws["A2"] = f"Erstellt: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws["A2"].font = Font(italic=True, color="666666")

        # Zusammenfassung
        row = 4
        if summary.get("text"):
            ws[f"A{row}"] = "Zusammenfassung"
            ws[f"A{row}"].font = self.SUBTITLE_FONT
            row += 1

            # Text umbrechen
            ws[f"A{row}"] = summary["text"]
            ws[f"A{row}"].alignment = Alignment(wrap_text=True)
            ws.merge_cells(f"A{row}:F{row}")
            ws.row_dimensions[row].height = 100
            row += 2

        # Key Metrics
        if summary.get("metrics"):
            ws[f"A{row}"] = "Kennzahlen"
            ws[f"A{row}"].font = self.SUBTITLE_FONT
            row += 1

            for key, value in summary["metrics"].items():
                ws[f"A{row}"] = key
                ws[f"B{row}"] = value
                row += 1

        # Spaltenbreiten anpassen
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

    def _create_data_sheet(
        self,
        wb: "Workbook",
        name: str,
        data: Dict[str, Any]
    ):
        """Erstellt ein Daten-Sheet."""
        ws = wb.create_sheet(name[:31])  # Excel max 31 chars

        headers = data.get("headers", [])
        rows = data.get("rows", [])

        # Header schreiben
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        # Daten schreiben
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Zahlen rechtsbündig
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")

                # Währung formatieren
                if isinstance(value, float) and data.get("currency_columns"):
                    if col_idx in data["currency_columns"]:
                        cell.number_format = '#,##0.00 €'

        # Spaltenbreiten anpassen
        for col in range(1, len(headers) + 1):
            max_length = max(
                len(str(headers[col - 1])),
                max((len(str(row[col - 1])) for row in rows if col <= len(row)), default=0)
            )
            ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 50)

        # Filter hinzufuegen
        if headers:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

        # Tabelle einfrieren
        ws.freeze_panes = "A2"

    def _add_chart(self, wb: "Workbook", config: Dict[str, Any]):
        """Fuegt einen Chart hinzu."""
        chart_type = config.get("type", "bar")
        sheet_name = config.get("sheet", "Zusammenfassung")
        data_range = config.get("data_range", {})

        if sheet_name not in wb.sheetnames:
            return

        ws = wb[sheet_name]

        # Chart erstellen
        if chart_type == "bar":
            chart = BarChart()
        elif chart_type == "pie":
            chart = PieChart()
        elif chart_type == "line":
            chart = LineChart()
        else:
            chart = BarChart()

        chart.title = config.get("title", "")
        chart.style = 10

        # Daten referenzieren
        if data_range:
            data = Reference(
                ws,
                min_col=data_range.get("min_col", 2),
                min_row=data_range.get("min_row", 1),
                max_col=data_range.get("max_col", 2),
                max_row=data_range.get("max_row", 10)
            )
            categories = Reference(
                ws,
                min_col=data_range.get("cat_col", 1),
                min_row=data_range.get("min_row", 2),
                max_row=data_range.get("max_row", 10)
            )
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)

        # Position
        position = config.get("position", "H2")
        ws.add_chart(chart, position)

    def create_customer_report(
        self,
        customer_name: str,
        summary: str,
        documents: List[Dict[str, Any]],
        metrics: Dict[str, Any],
        output_path: Optional[Path] = None
    ) -> bytes:
        """
        Erstellt einen Kunden-Report.

        Args:
            customer_name: Kundenname
            summary: LLM-generierte Zusammenfassung
            documents: Liste der Dokumente
            metrics: Kennzahlen
            output_path: Optionaler Speicherpfad

        Returns:
            Excel-Datei als Bytes
        """
        data = {
            "summary": {
                "text": summary,
                "metrics": metrics
            },
            "sheets": {
                "Dokumente": {
                    "headers": ["Datum", "Typ", "Titel", "Status"],
                    "rows": [
                        [
                            d.get("date", ""),
                            d.get("type", ""),
                            d.get("title", ""),
                            d.get("status", "")
                        ]
                        for d in documents
                    ]
                }
            },
            "charts": []
        }

        # Dokumenttyp-Verteilung als Chart
        type_counts = {}
        for doc in documents:
            doc_type = doc.get("type", "Unbekannt")
            type_counts[doc_type] = type_counts.get(doc_type, 0) + 1

        if type_counts:
            data["sheets"]["Dokumenttypen"] = {
                "headers": ["Typ", "Anzahl"],
                "rows": [[k, v] for k, v in type_counts.items()]
            }
            data["charts"].append({
                "type": "pie",
                "title": "Dokumenttypen",
                "sheet": "Dokumenttypen",
                "data_range": {
                    "min_col": 2,
                    "max_col": 2,
                    "min_row": 1,
                    "max_row": len(type_counts) + 1,
                    "cat_col": 1
                },
                "position": "D2"
            })

        return self.create_report(
            title=f"Kundenreport: {customer_name}",
            data=data,
            output_path=output_path
        )

    def create_supplier_report(
        self,
        supplier_name: str,
        summary: str,
        invoices: List[Dict[str, Any]],
        contracts: List[Dict[str, Any]],
        metrics: Dict[str, Any],
        output_path: Optional[Path] = None
    ) -> bytes:
        """
        Erstellt einen Lieferanten-Report.

        Args:
            supplier_name: Lieferantenname
            summary: LLM-generierte Zusammenfassung
            invoices: Rechnungsliste
            contracts: Vertragsliste
            metrics: Kennzahlen
            output_path: Optionaler Speicherpfad

        Returns:
            Excel-Datei als Bytes
        """
        data = {
            "summary": {
                "text": summary,
                "metrics": metrics
            },
            "sheets": {
                "Rechnungen": {
                    "headers": ["Rechnungsnr.", "Datum", "Betrag", "Status", "Faelligkeit"],
                    "rows": [
                        [
                            inv.get("number", ""),
                            inv.get("date", ""),
                            inv.get("amount", 0),
                            inv.get("status", ""),
                            inv.get("due_date", "")
                        ]
                        for inv in invoices
                    ],
                    "currency_columns": [3]
                },
                "Vertraege": {
                    "headers": ["Vertragsnr.", "Titel", "Beginn", "Ende", "Wert"],
                    "rows": [
                        [
                            c.get("number", ""),
                            c.get("title", ""),
                            c.get("start_date", ""),
                            c.get("end_date", ""),
                            c.get("value", 0)
                        ]
                        for c in contracts
                    ],
                    "currency_columns": [5]
                }
            },
            "charts": []
        }

        return self.create_report(
            title=f"Lieferantenreport: {supplier_name}",
            data=data,
            output_path=output_path
        )


# Singleton
_excel_generator: Optional[ExcelReportGenerator] = None


def get_excel_generator() -> ExcelReportGenerator:
    """Gibt ExcelReportGenerator Singleton zurueck."""
    global _excel_generator
    if _excel_generator is None:
        _excel_generator = ExcelReportGenerator()
    return _excel_generator
