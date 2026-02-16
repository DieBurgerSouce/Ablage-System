# -*- coding: utf-8 -*-
"""
Export Service.

Exportiert extrahierte Dokumentdaten in verschiedene Formate:
- CSV (für allgemeine Datenverarbeitung)
- Excel (.xlsx) mit Formatierung (für Buchhaltung)

Feinpoliert und durchdacht - Deutsche Dokumente mit hoechster Genauigkeit.
"""

import csv
import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Dict, List, Optional

import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = structlog.get_logger(__name__)


# =============================================================================
# COLUMN DEFINITIONS
# =============================================================================

INVOICE_COLUMNS = [
    ("Rechnungsnummer", "invoice_number"),
    ("Rechnungsdatum", "invoice_date"),
    ("Fälligkeitsdatum", "due_date"),
    ("Absender (Firma)", "sender_company"),
    ("Absender (Strasse)", "sender_street"),
    ("Absender (PLZ)", "sender_zip"),
    ("Absender (Ort)", "sender_city"),
    ("IBAN", "sender_iban"),
    ("USt-IdNr", "sender_vat_id"),
    ("Empfänger (Firma)", "recipient_company"),
    ("Kundennummer", "customer_number"),
    ("Bestellnummer", "order_number"),
    ("Nettobetrag", "net_amount"),
    ("MwSt-Satz (%)", "vat_rate"),
    ("MwSt-Betrag", "vat_amount"),
    ("Bruttobetrag", "gross_amount"),
    ("Währung", "currency"),
    ("Skonto (%)", "discount_percent"),
    ("Skonto-Frist", "discount_due_date"),
    ("Zahlungsziel (Tage)", "payment_days"),
    ("Konfidenz", "confidence"),
    ("Dokument-ID", "document_id"),
    ("Dateiname", "filename"),
]

ORDER_COLUMNS = [
    ("Bestellnummer", "order_number"),
    ("Bestelldatum", "order_date"),
    ("Liefertermin", "delivery_date"),
    ("Besteller (Firma)", "orderer_company"),
    ("Lieferant (Firma)", "supplier_company"),
    ("Gesamtbetrag", "total_amount"),
    ("Währung", "currency"),
    ("Konfidenz", "confidence"),
    ("Dokument-ID", "document_id"),
    ("Dateiname", "filename"),
]

CONTRACT_COLUMNS = [
    ("Vertragsnummer", "contract_number"),
    ("Vertragsdatum", "contract_date"),
    ("Vertragsbeginn", "start_date"),
    ("Vertragsende", "end_date"),
    ("Laufzeit", "duration"),
    ("Kündigungsfrist", "notice_period"),
    ("Partei A (Firma)", "party_a_company"),
    ("Partei B (Firma)", "party_b_company"),
    ("Vertragswert", "contract_value"),
    ("Monatlicher Wert", "monthly_value"),
    ("Vertragstyp", "contract_type"),
    ("Konfidenz", "confidence"),
    ("Dokument-ID", "document_id"),
    ("Dateiname", "filename"),
]


# =============================================================================
# DATA EXTRACTION HELPERS
# =============================================================================

def _extract_invoice_row(doc: Dict[str, Any]) -> Dict[str, Any]:
    """Extrahiert Rechnungsdaten für Export-Zeile."""
    extracted = doc.get("extracted_data", {}) or {}
    invoice = extracted.get("invoice", {}) or {}
    sender = invoice.get("sender", {}) or {}
    recipient = invoice.get("recipient", {}) or {}
    sender_bank = invoice.get("sender_bank", {}) or {}
    classification = extracted.get("classification", {}) or {}

    return {
        "invoice_number": invoice.get("invoice_number"),
        "invoice_date": invoice.get("invoice_date"),
        "due_date": invoice.get("due_date"),
        "sender_company": sender.get("company"),
        "sender_street": sender.get("street"),
        "sender_zip": sender.get("zip_code"),
        "sender_city": sender.get("city"),
        "sender_iban": sender_bank.get("iban"),
        "sender_vat_id": invoice.get("sender_vat_id"),
        "recipient_company": recipient.get("company"),
        "customer_number": invoice.get("customer_number"),
        "order_number": invoice.get("order_number"),
        "net_amount": invoice.get("net_amount"),
        "vat_rate": invoice.get("vat_rate"),
        "vat_amount": invoice.get("vat_amount"),
        "gross_amount": invoice.get("gross_amount"),
        "currency": invoice.get("currency", "EUR"),
        "discount_percent": invoice.get("discount_percent"),
        "discount_due_date": invoice.get("discount_due_date"),
        "payment_days": invoice.get("payment_days"),
        "confidence": classification.get("confidence", 0.0),
        "document_id": str(doc.get("id", "")),
        "filename": doc.get("filename"),
    }


def _extract_order_row(doc: Dict[str, Any]) -> Dict[str, Any]:
    """Extrahiert Bestellungsdaten für Export-Zeile."""
    extracted = doc.get("extracted_data", {}) or {}
    order = extracted.get("order", {}) or {}
    orderer = order.get("orderer", {}) or {}
    supplier = order.get("supplier", {}) or {}
    classification = extracted.get("classification", {}) or {}

    return {
        "order_number": order.get("order_number"),
        "order_date": order.get("order_date"),
        "delivery_date": order.get("delivery_date"),
        "orderer_company": orderer.get("company"),
        "supplier_company": supplier.get("company"),
        "total_amount": order.get("total_amount"),
        "currency": order.get("currency", "EUR"),
        "confidence": classification.get("confidence", 0.0),
        "document_id": str(doc.get("id", "")),
        "filename": doc.get("filename"),
    }


def _extract_contract_row(doc: Dict[str, Any]) -> Dict[str, Any]:
    """Extrahiert Vertragsdaten für Export-Zeile."""
    extracted = doc.get("extracted_data", {}) or {}
    contract = extracted.get("contract", {}) or {}
    party_a = contract.get("party_a", {}) or {}
    party_b = contract.get("party_b", {}) or {}
    classification = extracted.get("classification", {}) or {}

    return {
        "contract_number": contract.get("contract_number"),
        "contract_date": contract.get("contract_date"),
        "start_date": contract.get("start_date"),
        "end_date": contract.get("end_date"),
        "duration": contract.get("duration"),
        "notice_period": contract.get("notice_period"),
        "party_a_company": party_a.get("company"),
        "party_b_company": party_b.get("company"),
        "contract_value": contract.get("contract_value"),
        "monthly_value": contract.get("monthly_value"),
        "contract_type": contract.get("contract_type"),
        "confidence": classification.get("confidence", 0.0),
        "document_id": str(doc.get("id", "")),
        "filename": doc.get("filename"),
    }


def _format_value(value: object) -> str:
    """Formatiert einen Wert für CSV-Export."""
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat() if isinstance(value, datetime) else str(value)
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, float):
        return f"{value:.2f}"
    return str(value)


def _format_decimal_german(value: object) -> Optional[float]:
    """Konvertiert zu float für Excel-Formatierung."""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


# =============================================================================
# CSV EXPORT
# =============================================================================

def export_invoices_csv(documents: List[Dict[str, Any]]) -> str:
    """
    Exportiert Rechnungen als CSV.

    Args:
        documents: Liste von Document-Dicts mit extracted_data

    Returns:
        CSV-String mit UTF-8 BOM für Excel-Kompatibilität
    """
    output = io.StringIO()

    # UTF-8 BOM für Excel
    output.write('\ufeff')

    # Header
    headers = [col[0] for col in INVOICE_COLUMNS]
    writer = csv.DictWriter(
        output,
        fieldnames=[col[1] for col in INVOICE_COLUMNS],
        delimiter=';',  # Semikolon für deutsche Excel-Versionen
        quoting=csv.QUOTE_MINIMAL
    )

    # Write header with German names
    writer.writerow({col[1]: col[0] for col in INVOICE_COLUMNS})

    # Write data rows
    for doc in documents:
        row = _extract_invoice_row(doc)
        # Format all values
        formatted_row = {k: _format_value(v) for k, v in row.items()}
        writer.writerow(formatted_row)

    logger.info(
        "export_invoices_csv_complete",
        count=len(documents)
    )

    return output.getvalue()


def export_orders_csv(documents: List[Dict[str, Any]]) -> str:
    """Exportiert Bestellungen als CSV."""
    output = io.StringIO()
    output.write('\ufeff')

    writer = csv.DictWriter(
        output,
        fieldnames=[col[1] for col in ORDER_COLUMNS],
        delimiter=';',
        quoting=csv.QUOTE_MINIMAL
    )

    writer.writerow({col[1]: col[0] for col in ORDER_COLUMNS})

    for doc in documents:
        row = _extract_order_row(doc)
        formatted_row = {k: _format_value(v) for k, v in row.items()}
        writer.writerow(formatted_row)

    logger.info("export_orders_csv_complete", count=len(documents))

    return output.getvalue()


def export_contracts_csv(documents: List[Dict[str, Any]]) -> str:
    """Exportiert Verträge als CSV."""
    output = io.StringIO()
    output.write('\ufeff')

    writer = csv.DictWriter(
        output,
        fieldnames=[col[1] for col in CONTRACT_COLUMNS],
        delimiter=';',
        quoting=csv.QUOTE_MINIMAL
    )

    writer.writerow({col[1]: col[0] for col in CONTRACT_COLUMNS})

    for doc in documents:
        row = _extract_contract_row(doc)
        formatted_row = {k: _format_value(v) for k, v in row.items()}
        writer.writerow(formatted_row)

    logger.info("export_contracts_csv_complete", count=len(documents))

    return output.getvalue()


# =============================================================================
# EXCEL EXPORT
# =============================================================================

def _apply_excel_styling(ws, columns: List[tuple], row_count: int) -> None:
    """Wendet professionelles Styling auf Excel-Worksheet an."""
    # Header-Style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Rahmen
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header formatieren
    for col_idx, (header_name, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Daten-Zeilen formatieren
    data_alignment = Alignment(vertical="center")
    currency_format = '#,##0.00 "EUR"'
    percent_format = '0.00%'
    date_format = 'DD.MM.YYYY'

    for row_idx in range(2, row_count + 2):
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = data_alignment

    # Spaltenbreiten automatisch anpassen
    for col_idx, (header_name, field_name) in enumerate(columns, 1):
        column_letter = get_column_letter(col_idx)

        # Breite basierend auf Header-Name oder Feldtyp
        if "amount" in field_name.lower() or "betrag" in header_name.lower():
            ws.column_dimensions[column_letter].width = 15
        elif "date" in field_name.lower() or "datum" in header_name.lower():
            ws.column_dimensions[column_letter].width = 12
        elif "iban" in field_name.lower():
            ws.column_dimensions[column_letter].width = 25
        elif "company" in field_name.lower() or "firma" in header_name.lower():
            ws.column_dimensions[column_letter].width = 25
        elif "id" in field_name.lower():
            ws.column_dimensions[column_letter].width = 38
        else:
            ws.column_dimensions[column_letter].width = 15

    # Zeilenhöhe für Header
    ws.row_dimensions[1].height = 30


def export_invoices_excel(documents: List[Dict[str, Any]]) -> bytes:
    """
    Exportiert Rechnungen als formatiertes Excel (.xlsx).

    Args:
        documents: Liste von Document-Dicts mit extracted_data

    Returns:
        Excel-Datei als Bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Rechnungen"

    # Header schreiben
    for col_idx, (header_name, _) in enumerate(INVOICE_COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=header_name)

    # Daten schreiben
    for row_idx, doc in enumerate(documents, 2):
        row = _extract_invoice_row(doc)
        for col_idx, (_, field_name) in enumerate(INVOICE_COLUMNS, 1):
            value = row.get(field_name)

            # Zahlen direkt als float für Excel-Formatierung
            if field_name in ["net_amount", "vat_amount", "gross_amount", "discount_percent", "vat_rate"]:
                value = _format_decimal_german(value)
            elif field_name in ["confidence"]:
                value = _format_decimal_german(value)
                if value is not None:
                    value = value * 100  # Als Prozent

            ws.cell(row=row_idx, column=col_idx, value=value)

    # Styling anwenden
    _apply_excel_styling(ws, INVOICE_COLUMNS, len(documents))

    # Zahlenformate
    for row_idx in range(2, len(documents) + 2):
        for col_idx, (_, field_name) in enumerate(INVOICE_COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if field_name in ["net_amount", "vat_amount", "gross_amount"]:
                cell.number_format = '#,##0.00'
            elif field_name in ["discount_percent", "vat_rate", "confidence"]:
                cell.number_format = '0.00'

    # Autofilter aktivieren
    ws.auto_filter.ref = ws.dimensions

    # Erste Zeile fixieren
    ws.freeze_panes = "A2"

    # Als Bytes zurückgeben
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info(
        "export_invoices_excel_complete",
        count=len(documents)
    )

    return output.getvalue()


def export_orders_excel(documents: List[Dict[str, Any]]) -> bytes:
    """Exportiert Bestellungen als formatiertes Excel (.xlsx)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bestellungen"

    for col_idx, (header_name, _) in enumerate(ORDER_COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=header_name)

    for row_idx, doc in enumerate(documents, 2):
        row = _extract_order_row(doc)
        for col_idx, (_, field_name) in enumerate(ORDER_COLUMNS, 1):
            value = row.get(field_name)
            if field_name in ["total_amount"]:
                value = _format_decimal_german(value)
            elif field_name == "confidence":
                value = _format_decimal_german(value)
                if value is not None:
                    value = value * 100
            ws.cell(row=row_idx, column=col_idx, value=value)

    _apply_excel_styling(ws, ORDER_COLUMNS, len(documents))
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info("export_orders_excel_complete", count=len(documents))

    return output.getvalue()


def export_contracts_excel(documents: List[Dict[str, Any]]) -> bytes:
    """Exportiert Verträge als formatiertes Excel (.xlsx)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Verträge"

    for col_idx, (header_name, _) in enumerate(CONTRACT_COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=header_name)

    for row_idx, doc in enumerate(documents, 2):
        row = _extract_contract_row(doc)
        for col_idx, (_, field_name) in enumerate(CONTRACT_COLUMNS, 1):
            value = row.get(field_name)
            if field_name in ["contract_value", "monthly_value"]:
                value = _format_decimal_german(value)
            elif field_name == "confidence":
                value = _format_decimal_german(value)
                if value is not None:
                    value = value * 100
            ws.cell(row=row_idx, column=col_idx, value=value)

    _apply_excel_styling(ws, CONTRACT_COLUMNS, len(documents))
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info("export_contracts_excel_complete", count=len(documents))

    return output.getvalue()


# =============================================================================
# COMBINED EXPORT (alle Dokumenttypen in einer Datei)
# =============================================================================

def export_all_excel(
    invoices: List[Dict[str, Any]],
    orders: List[Dict[str, Any]],
    contracts: List[Dict[str, Any]]
) -> bytes:
    """
    Exportiert alle Dokumenttypen in eine Excel-Datei mit separaten Tabs.

    Args:
        invoices: Rechnungs-Dokumente
        orders: Bestellungs-Dokumente
        contracts: Vertrags-Dokumente

    Returns:
        Excel-Datei als Bytes
    """
    wb = Workbook()

    # Rechnungen
    ws_invoices = wb.active
    ws_invoices.title = "Rechnungen"

    for col_idx, (header_name, _) in enumerate(INVOICE_COLUMNS, 1):
        ws_invoices.cell(row=1, column=col_idx, value=header_name)

    for row_idx, doc in enumerate(invoices, 2):
        row = _extract_invoice_row(doc)
        for col_idx, (_, field_name) in enumerate(INVOICE_COLUMNS, 1):
            value = row.get(field_name)
            if field_name in ["net_amount", "vat_amount", "gross_amount", "discount_percent", "vat_rate", "confidence"]:
                value = _format_decimal_german(value)
            ws_invoices.cell(row=row_idx, column=col_idx, value=value)

    _apply_excel_styling(ws_invoices, INVOICE_COLUMNS, len(invoices))
    if invoices:
        ws_invoices.auto_filter.ref = ws_invoices.dimensions
        ws_invoices.freeze_panes = "A2"

    # Bestellungen
    ws_orders = wb.create_sheet("Bestellungen")

    for col_idx, (header_name, _) in enumerate(ORDER_COLUMNS, 1):
        ws_orders.cell(row=1, column=col_idx, value=header_name)

    for row_idx, doc in enumerate(orders, 2):
        row = _extract_order_row(doc)
        for col_idx, (_, field_name) in enumerate(ORDER_COLUMNS, 1):
            value = row.get(field_name)
            if field_name in ["total_amount", "confidence"]:
                value = _format_decimal_german(value)
            ws_orders.cell(row=row_idx, column=col_idx, value=value)

    _apply_excel_styling(ws_orders, ORDER_COLUMNS, len(orders))
    if orders:
        ws_orders.auto_filter.ref = ws_orders.dimensions
        ws_orders.freeze_panes = "A2"

    # Verträge
    ws_contracts = wb.create_sheet("Verträge")

    for col_idx, (header_name, _) in enumerate(CONTRACT_COLUMNS, 1):
        ws_contracts.cell(row=1, column=col_idx, value=header_name)

    for row_idx, doc in enumerate(contracts, 2):
        row = _extract_contract_row(doc)
        for col_idx, (_, field_name) in enumerate(CONTRACT_COLUMNS, 1):
            value = row.get(field_name)
            if field_name in ["contract_value", "monthly_value", "confidence"]:
                value = _format_decimal_german(value)
            ws_contracts.cell(row=row_idx, column=col_idx, value=value)

    _apply_excel_styling(ws_contracts, CONTRACT_COLUMNS, len(contracts))
    if contracts:
        ws_contracts.auto_filter.ref = ws_contracts.dimensions
        ws_contracts.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info(
        "export_all_excel_complete",
        invoices=len(invoices),
        orders=len(orders),
        contracts=len(contracts)
    )

    return output.getvalue()
