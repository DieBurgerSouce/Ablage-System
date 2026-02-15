# -*- coding: utf-8 -*-
"""
Ad-Hoc Report Service.

Kernlogik fuer Feature #12: Ad-Hoc Reporting.
Erstellt, validiert und fuehrt benutzerdefinierte Reports aus.

SICHERHEIT:
- Alle Spalten-/Tabellennamen werden gegen Whitelists validiert
- Keine rohen SQL-Injections moeglich (parametrisierte Queries)
- Company-Isolation via company_id Filter
"""

from __future__ import annotations

import csv
import io
import re
import time
import uuid
from datetime import datetime, timezone
from decimal import Decimal
from typing import Dict, List, Optional, Tuple, Union

import structlog
from sqlalchemy import and_, case, cast, desc, func, or_, select, String, Float, Integer
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.safe_errors import safe_error_detail, safe_error_log
from app.db.models import (
    Base,
    BusinessEntity,
    BankTransaction,
    Document,
)
from app.db.models_adhoc_reporting import (
    AdHocReport,
    AggregationType,
    DataSourceType,
    ExportFormat,
    ReportExecutionLog,
)

logger = structlog.get_logger(__name__)

# =============================================================================
# TYPE ALIASES (no Any)
# =============================================================================

FilterValue = Union[str, int, float, bool, List[str], List[int], None]
ColumnDef = Dict[str, Union[str, None]]
RowData = Dict[str, Union[str, int, float, bool, None, Decimal]]


# =============================================================================
# WHITELISTED COLUMN REGISTRY
# =============================================================================

# Each data source maps to: (SQLAlchemy model, dict of allowed_column_name -> column attribute)
# SECURITY: Only whitelisted columns can be queried. This prevents SQL injection.

_COLUMN_REGISTRY: Dict[str, Dict[str, object]] = {}


def _build_column_registry() -> Dict[str, Dict[str, object]]:
    """Builds the whitelisted column registry lazily.

    Returns mapping of data_source -> {column_name: sqlalchemy_column}.
    Uses deferred import for satellite models.
    """
    if _COLUMN_REGISTRY:
        return _COLUMN_REGISTRY

    # --- DOCUMENTS ---
    _COLUMN_REGISTRY["documents"] = {
        "id": Document.id,
        "filename": Document.filename,
        "original_filename": Document.original_filename,
        "file_size": Document.file_size,
        "mime_type": Document.mime_type,
        "document_type": Document.document_type,
        "status": Document.status,
        "page_count": Document.page_count,
        "ocr_backend_used": Document.ocr_backend_used,
        "ocr_confidence": Document.ocr_confidence,
        "processing_duration_ms": Document.processing_duration_ms,
        "has_umlauts": Document.has_umlauts,
        "german_validation_score": Document.german_validation_score,
        "detected_language": Document.detected_language,
        "upload_date": Document.upload_date,
        "processed_date": Document.processed_date,
        "created_at": Document.created_at,
        "updated_at": Document.updated_at,
        "company_id": Document.company_id,
        "business_entity_id": Document.business_entity_id,
        "chain_id": Document.chain_id,
        "chain_position": Document.chain_position,
    }

    # --- ENTITIES ---
    _COLUMN_REGISTRY["entities"] = {
        "id": BusinessEntity.id,
        "entity_type": BusinessEntity.entity_type,
        "name": BusinessEntity.name,
        "display_name": BusinessEntity.display_name,
        "vat_id": BusinessEntity.vat_id,
        "iban": BusinessEntity.iban,
        "postal_code": BusinessEntity.postal_code,
        "city": BusinessEntity.city,
        "country": BusinessEntity.country,
        "document_count": BusinessEntity.document_count,
        "total_invoice_amount": BusinessEntity.total_invoice_amount,
        "currency": BusinessEntity.currency,
        "is_active": BusinessEntity.is_active,
        "verified": BusinessEntity.verified,
        "confidence_score": BusinessEntity.confidence_score,
        "risk_score": BusinessEntity.risk_score,
        "payment_behavior_score": BusinessEntity.payment_behavior_score,
        "first_document_date": BusinessEntity.first_document_date,
        "last_document_date": BusinessEntity.last_document_date,
        "created_at": BusinessEntity.created_at,
        "updated_at": BusinessEntity.updated_at,
    }

    # --- TRANSACTIONS ---
    _COLUMN_REGISTRY["transactions"] = {
        "id": BankTransaction.id,
        "transaction_id": BankTransaction.transaction_id,
        "booking_date": BankTransaction.booking_date,
        "value_date": BankTransaction.value_date,
        "amount": BankTransaction.amount,
        "currency": BankTransaction.currency,
        "counterparty_name": BankTransaction.counterparty_name,
        "reference_text": BankTransaction.reference_text,
        "transaction_type": BankTransaction.transaction_type,
        "booking_text": BankTransaction.booking_text,
        "reconciliation_status": BankTransaction.reconciliation_status,
        "match_confidence": BankTransaction.match_confidence,
        "is_partial_payment": BankTransaction.is_partial_payment,
        "created_at": BankTransaction.created_at,
    }

    # --- INVOICES (satellite model, deferred import) ---
    try:
        from app.db.models_invoice import Invoice
        _COLUMN_REGISTRY["invoices"] = {
            "id": Invoice.id,
            "invoice_number": Invoice.invoice_number,
            "invoice_date": Invoice.invoice_date,
            "due_date": Invoice.due_date,
            "subtotal": Invoice.subtotal,
            "tax_amount": Invoice.tax_amount,
            "total_amount": Invoice.total_amount,
            "currency": Invoice.currency,
            "status": Invoice.status,
            "payment_date": Invoice.payment_date,
            "notes": Invoice.notes,
            "created_at": Invoice.created_at,
        }
    except ImportError:
        logger.warning("invoice_model_not_available", msg="Invoice model konnte nicht geladen werden")
        _COLUMN_REGISTRY["invoices"] = {}

    return _COLUMN_REGISTRY


def _get_model_for_source(source: str) -> Optional[type]:
    """Returns the SQLAlchemy model class for a given data source.

    SECURITY: Only whitelisted sources are allowed.
    """
    source_model_map: Dict[str, type] = {
        "documents": Document,
        "entities": BusinessEntity,
        "transactions": BankTransaction,
    }

    # Deferred import for satellite models
    if source == "invoices":
        try:
            from app.db.models_invoice import Invoice
            return Invoice
        except ImportError:
            return None

    return source_model_map.get(source)


# =============================================================================
# FILTER OPERATORS
# =============================================================================

# Whitelist of allowed filter operators
ALLOWED_OPERATORS = frozenset({
    "eq", "ne", "gt", "gte", "lt", "lte",
    "contains", "starts_with", "ends_with",
    "in", "not_in", "is_null", "is_not_null",
    "between",
})

# Regex for safe column names (letters, digits, underscore only)
_SAFE_COLUMN_RE = re.compile(r"^[a-zA-Z_][a-zA-Z0-9_]{0,63}$")


def _validate_column_name(name: str) -> bool:
    """Validates a column name against the safe pattern.

    SECURITY: Prevents SQL injection through column names.
    """
    return bool(_SAFE_COLUMN_RE.match(name))


# =============================================================================
# DATA SOURCE METADATA
# =============================================================================


def get_available_data_sources() -> List[Dict[str, str]]:
    """Gibt die verfuegbaren Datenquellen mit Beschreibung zurueck."""
    return [
        {
            "id": DataSourceType.INVOICES.value,
            "name": "Rechnungen",
            "description": "Rechnungsdaten mit Betraegen, Faelligkeiten und Status",
        },
        {
            "id": DataSourceType.DOCUMENTS.value,
            "name": "Dokumente",
            "description": "Alle hochgeladenen Dokumente mit OCR-Ergebnissen",
        },
        {
            "id": DataSourceType.ENTITIES.value,
            "name": "Geschaeftspartner",
            "description": "Kunden und Lieferanten mit Risiko-Scores",
        },
        {
            "id": DataSourceType.TRANSACTIONS.value,
            "name": "Kontobewegungen",
            "description": "Importierte Banktransaktionen mit Reconciliation-Status",
        },
    ]


def get_available_columns(data_source: str) -> List[Dict[str, str]]:
    """Gibt die verfuegbaren Spalten fuer eine Datenquelle zurueck.

    SECURITY: Returns only whitelisted columns.
    """
    registry = _build_column_registry()
    source_columns = registry.get(data_source, {})

    result: List[Dict[str, str]] = []
    for col_name, col_attr in source_columns.items():
        # Determine data type from SQLAlchemy column
        col_type = "string"
        if hasattr(col_attr, "type"):
            type_name = type(col_attr.type).__name__.lower()
            if type_name in ("integer", "biginteger"):
                col_type = "number"
            elif type_name in ("float", "numeric"):
                col_type = "number"
            elif type_name in ("datetime",):
                col_type = "datetime"
            elif type_name in ("date",):
                col_type = "date"
            elif type_name in ("boolean",):
                col_type = "boolean"

        result.append({
            "name": col_name,
            "data_type": col_type,
            "source": data_source,
        })

    return result


# =============================================================================
# SERVICE
# =============================================================================


class AdHocReportService:
    """Service fuer Ad-Hoc Report CRUD und Ausfuehrung."""

    # ------------------------------------------------------------------
    # VALIDATION
    # ------------------------------------------------------------------

    def validate_report_config(
        self,
        data_sources: List[Dict[str, str]],
        columns: List[ColumnDef],
        filters: List[Dict[str, FilterValue]],
        group_by: Optional[List[str]] = None,
        order_by: Optional[List[Dict[str, str]]] = None,
    ) -> List[str]:
        """Validiert die Report-Konfiguration gegen Whitelists.

        Returns:
            Leere Liste bei Erfolg, sonst Liste der Fehler.
        """
        errors: List[str] = []
        registry = _build_column_registry()

        if not data_sources:
            errors.append("Mindestens eine Datenquelle ist erforderlich")
            return errors

        # Validate data sources
        valid_sources = {ds.value for ds in DataSourceType}
        source_names: List[str] = []
        for ds_config in data_sources:
            source = ds_config.get("source", "")
            if source not in valid_sources:
                errors.append(f"Ungueltige Datenquelle: {source}")
            elif source not in registry or not registry[source]:
                errors.append(f"Datenquelle nicht verfuegbar: {source}")
            else:
                source_names.append(source)

        if errors:
            return errors

        # Validate columns
        if not columns:
            errors.append("Mindestens eine Spalte ist erforderlich")
            return errors

        for col in columns:
            col_name = col.get("name", "")
            col_source = col.get("source", source_names[0] if source_names else "")

            if not _validate_column_name(col_name):
                errors.append(f"Ungueltiger Spaltenname: {col_name}")
                continue

            if col_source not in source_names:
                errors.append(f"Spalte '{col_name}' referenziert unbekannte Quelle: {col_source}")
                continue

            source_cols = registry.get(col_source, {})
            if col_name not in source_cols:
                errors.append(f"Spalte '{col_name}' nicht in Quelle '{col_source}' gefunden")

            # Validate aggregation
            agg = col.get("aggregation")
            if agg:
                valid_aggs = {a.value for a in AggregationType}
                if agg not in valid_aggs:
                    errors.append(f"Ungueltige Aggregation '{agg}' fuer Spalte '{col_name}'")

        # Validate filters
        for flt in filters:
            flt_column = flt.get("column", "")
            flt_operator = flt.get("operator", "")
            flt_source = flt.get("source", source_names[0] if source_names else "")

            if not _validate_column_name(str(flt_column)):
                errors.append(f"Ungueltiger Filter-Spaltenname: {flt_column}")
                continue

            if flt_operator not in ALLOWED_OPERATORS:
                errors.append(f"Ungueltiger Operator: {flt_operator}")

            if isinstance(flt_source, str) and flt_source in registry:
                source_cols = registry[flt_source]
                if flt_column not in source_cols:
                    errors.append(
                        f"Filter-Spalte '{flt_column}' nicht in Quelle '{flt_source}' gefunden"
                    )

        # Validate group_by
        if group_by:
            for gb_col in group_by:
                if not _validate_column_name(gb_col):
                    errors.append(f"Ungueltiger GROUP BY Spaltenname: {gb_col}")

        # Validate order_by
        if order_by:
            for ob in order_by:
                ob_col = ob.get("column", "")
                ob_dir = ob.get("direction", "asc")
                if not _validate_column_name(ob_col):
                    errors.append(f"Ungueltiger ORDER BY Spaltenname: {ob_col}")
                if ob_dir not in ("asc", "desc"):
                    errors.append(f"Ungueltige Sortierrichtung: {ob_dir}")

        return errors

    # ------------------------------------------------------------------
    # CRUD
    # ------------------------------------------------------------------

    async def create_report(
        self,
        db: AsyncSession,
        company_id: uuid.UUID,
        user_id: uuid.UUID,
        name: str,
        data_sources: List[Dict[str, str]],
        columns: List[ColumnDef],
        filters: Optional[List[Dict[str, FilterValue]]] = None,
        group_by: Optional[List[str]] = None,
        order_by: Optional[List[Dict[str, str]]] = None,
        limit_rows: Optional[int] = None,
        chart_config: Optional[Dict[str, str]] = None,
        description: Optional[str] = None,
        is_template: bool = False,
    ) -> AdHocReport:
        """Erstellt einen neuen Ad-Hoc Report."""
        report = AdHocReport(
            company_id=company_id,
            created_by_user_id=user_id,
            name=name,
            description=description,
            data_sources=data_sources,
            columns=columns,
            filters=filters or [],
            group_by=group_by,
            order_by=order_by,
            limit_rows=limit_rows,
            chart_config=chart_config,
            is_template=is_template,
        )
        db.add(report)
        await db.commit()
        await db.refresh(report)

        logger.info(
            "adhoc_report_created",
            report_id=str(report.id),
            name=name,
            sources=[ds.get("source") for ds in data_sources],
        )
        return report

    async def get_report(
        self,
        db: AsyncSession,
        report_id: uuid.UUID,
        company_id: uuid.UUID,
    ) -> Optional[AdHocReport]:
        """Laedt einen Report nach ID mit Company-Isolation."""
        result = await db.execute(
            select(AdHocReport).where(
                and_(
                    AdHocReport.id == report_id,
                    AdHocReport.company_id == company_id,
                    AdHocReport.deleted_at.is_(None),
                )
            )
        )
        return result.scalar_one_or_none()

    async def get_user_reports(
        self,
        db: AsyncSession,
        company_id: uuid.UUID,
        user_id: uuid.UUID,
        limit: int = 100,
        offset: int = 0,
    ) -> List[AdHocReport]:
        """Listet Reports eines Benutzers auf."""
        result = await db.execute(
            select(AdHocReport)
            .where(
                and_(
                    AdHocReport.company_id == company_id,
                    AdHocReport.created_by_user_id == user_id,
                    AdHocReport.deleted_at.is_(None),
                )
            )
            .order_by(desc(AdHocReport.updated_at))
            .limit(limit)
            .offset(offset)
        )
        return list(result.scalars().all())

    async def get_shared_reports(
        self,
        db: AsyncSession,
        company_id: uuid.UUID,
        limit: int = 100,
        offset: int = 0,
    ) -> List[AdHocReport]:
        """Listet oeffentlich geteilte Reports auf."""
        result = await db.execute(
            select(AdHocReport)
            .where(
                and_(
                    AdHocReport.company_id == company_id,
                    AdHocReport.is_shared == True,  # noqa: E712
                    AdHocReport.deleted_at.is_(None),
                )
            )
            .order_by(desc(AdHocReport.updated_at))
            .limit(limit)
            .offset(offset)
        )
        return list(result.scalars().all())

    async def update_report(
        self,
        db: AsyncSession,
        report_id: uuid.UUID,
        company_id: uuid.UUID,
        user_id: uuid.UUID,
        **updates: object,
    ) -> Optional[AdHocReport]:
        """Aktualisiert einen Report (nur Besitzer)."""
        report = await self.get_report(db, report_id, company_id)
        if not report:
            return None
        if report.created_by_user_id != user_id:
            return None

        allowed_fields = {
            "name", "description", "data_sources", "columns", "filters",
            "group_by", "order_by", "limit_rows", "chart_config",
            "is_template", "is_shared", "shared_with_users",
        }

        for key, value in updates.items():
            if key in allowed_fields:
                setattr(report, key, value)

        await db.commit()
        await db.refresh(report)
        return report

    async def delete_report(
        self,
        db: AsyncSession,
        report_id: uuid.UUID,
        company_id: uuid.UUID,
        user_id: uuid.UUID,
    ) -> bool:
        """Soft-Delete eines Reports (nur Besitzer)."""
        report = await self.get_report(db, report_id, company_id)
        if not report:
            return False
        if report.created_by_user_id != user_id:
            return False

        report.deleted_at = datetime.now(timezone.utc)
        await db.commit()

        logger.info("adhoc_report_deleted", report_id=str(report_id))
        return True

    async def share_report(
        self,
        db: AsyncSession,
        report_id: uuid.UUID,
        company_id: uuid.UUID,
        user_id: uuid.UUID,
        share_with_user_ids: List[str],
    ) -> Optional[AdHocReport]:
        """Teilt einen Report mit bestimmten Benutzern."""
        report = await self.get_report(db, report_id, company_id)
        if not report:
            return None
        if report.created_by_user_id != user_id:
            return None

        report.is_shared = True
        report.shared_with_users = share_with_user_ids
        await db.commit()
        await db.refresh(report)

        logger.info(
            "adhoc_report_shared",
            report_id=str(report_id),
            shared_with_count=len(share_with_user_ids),
        )
        return report

    async def duplicate_report(
        self,
        db: AsyncSession,
        report_id: uuid.UUID,
        company_id: uuid.UUID,
        user_id: uuid.UUID,
    ) -> Optional[AdHocReport]:
        """Dupliziert einen Report."""
        original = await self.get_report(db, report_id, company_id)
        if not original:
            return None

        copy = AdHocReport(
            company_id=company_id,
            created_by_user_id=user_id,
            name=f"{original.name} (Kopie)",
            description=original.description,
            data_sources=original.data_sources,
            columns=original.columns,
            filters=original.filters,
            group_by=original.group_by,
            order_by=original.order_by,
            limit_rows=original.limit_rows,
            chart_config=original.chart_config,
            is_template=False,
            is_shared=False,
        )
        db.add(copy)
        await db.commit()
        await db.refresh(copy)

        logger.info(
            "adhoc_report_duplicated",
            original_id=str(report_id),
            copy_id=str(copy.id),
        )
        return copy

    # ------------------------------------------------------------------
    # QUERY BUILDER
    # ------------------------------------------------------------------

    def build_query(
        self,
        report: AdHocReport,
        company_id: uuid.UUID,
    ) -> Tuple[object, List[str]]:
        """Baut eine SQLAlchemy-Query aus der Report-Konfiguration.

        SECURITY:
        - Alle Spalten werden gegen das Whitelist-Registry validiert
        - Company-Isolation wird IMMER erzwungen
        - Keine Raw-SQL, nur SQLAlchemy-Ausdruecke

        Returns:
            Tuple of (sqlalchemy select statement, list of column aliases)
        """
        registry = _build_column_registry()

        data_sources = report.data_sources or []
        if not data_sources:
            raise ValueError("Keine Datenquellen konfiguriert")

        # Determine primary data source
        primary_source = data_sources[0].get("source", "")
        primary_model = _get_model_for_source(primary_source)
        if not primary_model:
            raise ValueError(f"Unbekannte Datenquelle: {primary_source}")

        # Build SELECT columns
        select_columns = []
        column_aliases: List[str] = []
        report_columns = report.columns or []

        has_aggregation = False

        for col_def in report_columns:
            col_name = col_def.get("name", "")
            col_source = col_def.get("source", primary_source)
            col_alias = col_def.get("alias", col_name)
            col_agg = col_def.get("aggregation")

            # SECURITY: Validate against whitelist
            if not _validate_column_name(col_name):
                raise ValueError(f"Ungueltiger Spaltenname: {col_name}")

            source_cols = registry.get(col_source, {})
            if col_name not in source_cols:
                raise ValueError(
                    f"Spalte '{col_name}' nicht in Quelle '{col_source}' erlaubt"
                )

            sa_column = source_cols[col_name]

            # Apply aggregation
            if col_agg:
                has_aggregation = True
                if col_agg == AggregationType.COUNT.value:
                    sa_column = func.count(sa_column)
                elif col_agg == AggregationType.SUM.value:
                    sa_column = func.sum(sa_column)
                elif col_agg == AggregationType.AVG.value:
                    sa_column = func.avg(sa_column)
                elif col_agg == AggregationType.MIN.value:
                    sa_column = func.min(sa_column)
                elif col_agg == AggregationType.MAX.value:
                    sa_column = func.max(sa_column)
                elif col_agg == AggregationType.DISTINCT_COUNT.value:
                    sa_column = func.count(func.distinct(sa_column))

            select_columns.append(sa_column.label(col_alias))
            column_aliases.append(col_alias)

        if not select_columns:
            raise ValueError("Keine gueltigen Spalten konfiguriert")

        # Build base query
        stmt = select(*select_columns).select_from(primary_model)

        # Apply JOINs for additional data sources
        for ds_config in data_sources[1:]:
            join_source = ds_config.get("source", "")
            join_model = _get_model_for_source(join_source)
            if not join_model:
                continue

            join_on = ds_config.get("join_on", {})
            left_col_name = join_on.get("left", "")
            right_col_name = join_on.get("right", "")

            if left_col_name and right_col_name:
                left_cols = registry.get(primary_source, {})
                right_cols = registry.get(join_source, {})

                if left_col_name in left_cols and right_col_name in right_cols:
                    left_attr = left_cols[left_col_name]
                    right_attr = right_cols[right_col_name]
                    stmt = stmt.outerjoin(join_model, left_attr == right_attr)
                else:
                    logger.warning(
                        "adhoc_join_skipped",
                        left=left_col_name,
                        right=right_col_name,
                        reason="Spalten nicht in Whitelist",
                    )

        # SECURITY: Always enforce company isolation
        company_col = self._get_company_column(primary_model, primary_source, registry)
        if company_col is not None:
            stmt = stmt.where(company_col == company_id)

        # Apply filters
        filter_conditions = self._build_filter_conditions(
            report.filters or [],
            primary_source,
            registry,
        )
        if filter_conditions:
            stmt = stmt.where(and_(*filter_conditions))

        # Apply GROUP BY
        if report.group_by and has_aggregation:
            for gb_col_name in report.group_by:
                if not _validate_column_name(gb_col_name):
                    continue
                source_cols = registry.get(primary_source, {})
                if gb_col_name in source_cols:
                    stmt = stmt.group_by(source_cols[gb_col_name])

        # Apply ORDER BY
        if report.order_by:
            for ob in report.order_by:
                ob_col_name = ob.get("column", "")
                ob_direction = ob.get("direction", "asc")

                if not _validate_column_name(ob_col_name):
                    continue

                # Try to find the column in any source
                ob_col_attr = None
                for src_name in [primary_source] + [
                    ds.get("source", "") for ds in data_sources[1:]
                ]:
                    src_cols = registry.get(src_name, {})
                    if ob_col_name in src_cols:
                        ob_col_attr = src_cols[ob_col_name]
                        break

                if ob_col_attr is not None:
                    if ob_direction == "desc":
                        stmt = stmt.order_by(desc(ob_col_attr))
                    else:
                        stmt = stmt.order_by(ob_col_attr)

        # Apply LIMIT
        if report.limit_rows and report.limit_rows > 0:
            stmt = stmt.limit(min(report.limit_rows, 10000))
        else:
            stmt = stmt.limit(10000)  # Safety limit

        return stmt, column_aliases

    def _get_company_column(
        self,
        model: type,
        source_name: str,
        registry: Dict[str, Dict[str, object]],
    ) -> object:
        """Returns the company_id column for a model, if available."""
        source_cols = registry.get(source_name, {})
        if "company_id" in source_cols:
            return source_cols["company_id"]
        return None

    def _build_filter_conditions(
        self,
        filters: List[Dict[str, FilterValue]],
        primary_source: str,
        registry: Dict[str, Dict[str, object]],
    ) -> List[object]:
        """Builds SQLAlchemy filter conditions from the report filters.

        SECURITY: All column names and operators are validated against whitelists.
        """
        conditions = []

        for flt in filters:
            col_name = str(flt.get("column", ""))
            operator = str(flt.get("operator", "eq"))
            value = flt.get("value")
            source = str(flt.get("source", primary_source))

            if not _validate_column_name(col_name):
                continue
            if operator not in ALLOWED_OPERATORS:
                continue

            source_cols = registry.get(source, {})
            if col_name not in source_cols:
                continue

            sa_col = source_cols[col_name]

            try:
                condition = self._apply_operator(sa_col, operator, value)
                if condition is not None:
                    conditions.append(condition)
            except (ValueError, TypeError) as exc:
                logger.warning(
                    "adhoc_filter_skipped",
                    column=col_name,
                    operator=operator,
                    reason=str(exc),
                )

        return conditions

    def _apply_operator(
        self,
        column: object,
        operator: str,
        value: FilterValue,
    ) -> object:
        """Applies a filter operator to a SQLAlchemy column expression.

        SECURITY: Only whitelisted operators are supported.
        """
        if operator == "eq":
            return column == value
        elif operator == "ne":
            return column != value
        elif operator == "gt":
            return column > value
        elif operator == "gte":
            return column >= value
        elif operator == "lt":
            return column < value
        elif operator == "lte":
            return column <= value
        elif operator == "contains":
            return cast(column, String).ilike(f"%{value}%")
        elif operator == "starts_with":
            return cast(column, String).ilike(f"{value}%")
        elif operator == "ends_with":
            return cast(column, String).ilike(f"%{value}")
        elif operator == "in":
            if isinstance(value, list):
                return column.in_(value)
            return None
        elif operator == "not_in":
            if isinstance(value, list):
                return column.notin_(value)
            return None
        elif operator == "is_null":
            return column.is_(None)
        elif operator == "is_not_null":
            return column.isnot(None)
        elif operator == "between":
            if isinstance(value, list) and len(value) == 2:
                return column.between(value[0], value[1])
            return None
        return None

    # ------------------------------------------------------------------
    # EXECUTION
    # ------------------------------------------------------------------

    async def execute_report(
        self,
        db: AsyncSession,
        report_id: uuid.UUID,
        company_id: uuid.UUID,
        user_id: Optional[uuid.UUID] = None,
        export_format: Optional[str] = None,
    ) -> Dict[str, Union[List[RowData], List[str], int, str]]:
        """Fuehrt einen Report aus und gibt die Ergebnisse zurueck.

        Returns:
            Dict mit keys: rows, columns, row_count, execution_time_ms
        """
        report = await self.get_report(db, report_id, company_id)
        if not report:
            raise ValueError("Report nicht gefunden")

        start_time = time.monotonic()

        try:
            stmt, column_aliases = self.build_query(report, company_id)
            result = await db.execute(stmt)
            raw_rows = result.fetchall()
        except Exception as exc:
            # Log execution failure
            log_entry = ReportExecutionLog(
                report_id=report_id,
                company_id=company_id,
                executed_by_user_id=user_id,
                export_format=export_format,
                row_count=0,
                execution_time_ms=int((time.monotonic() - start_time) * 1000),
                error_message=safe_error_detail(exc, "Report"),
            )
            db.add(log_entry)
            await db.commit()
            raise

        elapsed_ms = int((time.monotonic() - start_time) * 1000)

        # Convert rows to dicts
        rows: List[RowData] = []
        for raw_row in raw_rows:
            row_dict: RowData = {}
            for idx, alias in enumerate(column_aliases):
                val = raw_row[idx]
                # Convert Decimal to float for JSON serialization
                if isinstance(val, Decimal):
                    val = float(val)
                # Convert datetime to ISO string
                if isinstance(val, datetime):
                    val = val.isoformat()
                # Convert uuid to string
                if isinstance(val, uuid.UUID):
                    val = str(val)
                row_dict[alias] = val
            rows.append(row_dict)

        # Update report execution stats
        report.last_executed_at = datetime.now(timezone.utc)
        report.execution_count = (report.execution_count or 0) + 1

        # Log execution
        log_entry = ReportExecutionLog(
            report_id=report_id,
            company_id=company_id,
            executed_by_user_id=user_id,
            export_format=export_format,
            row_count=len(rows),
            execution_time_ms=elapsed_ms,
        )
        db.add(log_entry)
        await db.commit()

        logger.info(
            "adhoc_report_executed",
            report_id=str(report_id),
            row_count=len(rows),
            execution_time_ms=elapsed_ms,
        )

        return {
            "rows": rows,
            "columns": column_aliases,
            "row_count": len(rows),
            "execution_time_ms": elapsed_ms,
        }

    # ------------------------------------------------------------------
    # EXPORT
    # ------------------------------------------------------------------

    def export_csv(
        self,
        rows: List[RowData],
        columns: List[str],
    ) -> bytes:
        """Generiert CSV-Bytes aus den Report-Daten."""
        output = io.StringIO()
        writer = csv.DictWriter(
            output,
            fieldnames=columns,
            extrasaction="ignore",
            delimiter=";",  # German Excel compatibility
        )
        writer.writeheader()
        for row in rows:
            # Convert all values to strings for CSV
            csv_row = {k: str(v) if v is not None else "" for k, v in row.items()}
            writer.writerow(csv_row)

        return output.getvalue().encode("utf-8-sig")  # BOM for Excel

    def export_excel(
        self,
        rows: List[RowData],
        columns: List[str],
        report_name: str = "Report",
    ) -> bytes:
        """Generiert Excel-Bytes aus den Report-Daten.

        Verwendet openpyxl fuer native XLSX-Generierung.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            logger.warning("openpyxl_not_available", msg="Fallback auf CSV")
            return self.export_csv(rows, columns)

        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = report_name[:31]  # Excel sheet name limit

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write headers
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = row_data.get(col_name)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Auto-adjust column widths
        for col_idx, col_name in enumerate(columns, 1):
            max_width = len(str(col_name))
            for row_idx in range(2, min(len(rows) + 2, 102)):  # Sample first 100 rows
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_width = max(max_width, len(str(cell_value)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
                max_width + 2, 50
            )

        # Freeze header row
        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
