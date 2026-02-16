# -*- coding: utf-8 -*-
"""
Ad-Hoc Report Builder Service.

Self-Service Reporting ohne IT-Hilfe:
- Daten frei kombinieren: Rechnungen x Lieferanten x Zeitraum x Kategorie
- Filter, Gruppierung, Aggregation
- Export: PDF, Excel, CSV
- Gespeicherte Reports teilen mit Kollegen
- Geplante Report-Ausführungen

SICHERHEIT:
- Alle Feldnamen werden gegen Whitelists validiert (SQL-Injection-Schutz)
- Alle Werte werden als Parameter gebunden (keine String-Konkatenation)
- Multi-Tenancy durch company_id-Filter in jeder Query

Feinpoliert und durchdacht - Enterprise-grade Ad-Hoc Reporting.
"""

import csv
import io
import os
import re
import time
import uuid
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Optional, Tuple, Union

import structlog
from sqlalchemy import text, select, update, delete, and_, or_, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.safe_errors import safe_error_log, safe_error_detail
from app.db.models_adhoc_report import (
    AdHocExportFormat,
    AdHocReport,
    AdHocReportExecution,
    AdHocReportShare,
    AggregationType,
    DataSourceType,
    ReportSchedule,
    ReportScheduleFrequency,
)

logger = structlog.get_logger(__name__)

# Typ-Alias (kein Any)
FilterValue = Union[str, int, float, bool, None]


class AdHocReportService:
    """Ad-Hoc Report Builder - Self-Service Reporting ohne IT-Hilfe.

    Daten frei kombinieren: Rechnungen x Lieferanten x Zeitraum x Kategorie.
    Filter, Gruppierung, Aggregation. Export: PDF, Excel, CSV.
    Gespeicherte Reports teilen mit Kollegen.
    """

    # =========================================================================
    # Datenquellen-Mapping (Tabelle, PK, Label)
    # =========================================================================

    DATA_SOURCES: Dict[str, Dict[str, object]] = {
        "invoices": {
            "table": "invoice_tracking",
            "key": "id",
            "label": "Rechnungen",
        },
        "documents": {
            "table": "documents",
            "key": "id",
            "label": "Dokumente",
        },
        "suppliers": {
            "table": "business_entities",
            "key": "id",
            "filter": {"entity_type": "supplier"},
            "label": "Lieferanten",
        },
        "customers": {
            "table": "business_entities",
            "key": "id",
            "filter": {"entity_type": "customer"},
            "label": "Kunden",
        },
        "transactions": {
            "table": "bank_transactions",
            "key": "id",
            "label": "Transaktionen",
        },
        "approvals": {
            "table": "approval_requests",
            "key": "id",
            "label": "Genehmigungen",
        },
        "workflows": {
            "table": "workflow_executions",
            "key": "id",
            "label": "Workflows",
        },
    }

    # =========================================================================
    # Erlaubte Felder pro Datenquelle (Whitelist für SQL-Injection-Schutz!)
    # =========================================================================

    AVAILABLE_FIELDS: Dict[str, List[str]] = {
        "invoices": [
            "invoice_number", "supplier_name", "net_amount", "gross_amount",
            "tax_amount", "status", "due_date", "paid_date", "payment_method",
            "created_at", "category",
        ],
        "documents": [
            "title", "document_type", "status", "ocr_status", "page_count",
            "file_size", "created_at", "updated_at",
        ],
        "suppliers": [
            "name", "vat_id", "city", "postal_code", "country",
            "risk_score", "created_at",
        ],
        "customers": [
            "name", "vat_id", "city", "postal_code", "country", "created_at",
        ],
        "transactions": [
            "amount", "currency", "booking_date", "value_date", "purpose",
            "transaction_type",
        ],
        "approvals": [
            "status", "priority", "requested_at", "decided_at",
            "approver_name", "document_type", "amount",
        ],
        "workflows": [
            "workflow_type", "status", "started_at", "completed_at",
            "duration_ms",
        ],
    }

    # Feld-Labels (Deutsch) und Datentypen
    FIELD_METADATA: Dict[str, Dict[str, Dict[str, str]]] = {
        "invoices": {
            "invoice_number": {"label": "Rechnungsnummer", "type": "string"},
            "supplier_name": {"label": "Lieferantenname", "type": "string"},
            "net_amount": {"label": "Nettobetrag", "type": "currency"},
            "gross_amount": {"label": "Bruttobetrag", "type": "currency"},
            "tax_amount": {"label": "Steuerbetrag", "type": "currency"},
            "status": {"label": "Status", "type": "string"},
            "due_date": {"label": "Fälligkeitsdatum", "type": "date"},
            "paid_date": {"label": "Bezahldatum", "type": "date"},
            "payment_method": {"label": "Zahlungsart", "type": "string"},
            "created_at": {"label": "Erstellt am", "type": "datetime"},
            "category": {"label": "Kategorie", "type": "string"},
        },
        "documents": {
            "title": {"label": "Titel", "type": "string"},
            "document_type": {"label": "Dokumenttyp", "type": "string"},
            "status": {"label": "Status", "type": "string"},
            "ocr_status": {"label": "OCR-Status", "type": "string"},
            "page_count": {"label": "Seitenzahl", "type": "number"},
            "file_size": {"label": "Dateigröße", "type": "number"},
            "created_at": {"label": "Erstellt am", "type": "datetime"},
            "updated_at": {"label": "Aktualisiert am", "type": "datetime"},
        },
        "suppliers": {
            "name": {"label": "Name", "type": "string"},
            "vat_id": {"label": "USt-IdNr.", "type": "string"},
            "city": {"label": "Stadt", "type": "string"},
            "postal_code": {"label": "PLZ", "type": "string"},
            "country": {"label": "Land", "type": "string"},
            "risk_score": {"label": "Risikobewertung", "type": "number"},
            "created_at": {"label": "Erstellt am", "type": "datetime"},
        },
        "customers": {
            "name": {"label": "Name", "type": "string"},
            "vat_id": {"label": "USt-IdNr.", "type": "string"},
            "city": {"label": "Stadt", "type": "string"},
            "postal_code": {"label": "PLZ", "type": "string"},
            "country": {"label": "Land", "type": "string"},
            "created_at": {"label": "Erstellt am", "type": "datetime"},
        },
        "transactions": {
            "amount": {"label": "Betrag", "type": "currency"},
            "currency": {"label": "Währung", "type": "string"},
            "booking_date": {"label": "Buchungsdatum", "type": "date"},
            "value_date": {"label": "Wertstellungsdatum", "type": "date"},
            "purpose": {"label": "Verwendungszweck", "type": "string"},
            "transaction_type": {"label": "Transaktionstyp", "type": "string"},
        },
        "approvals": {
            "status": {"label": "Status", "type": "string"},
            "priority": {"label": "Priorität", "type": "string"},
            "requested_at": {"label": "Angefragt am", "type": "datetime"},
            "decided_at": {"label": "Entschieden am", "type": "datetime"},
            "approver_name": {"label": "Genehmiger", "type": "string"},
            "document_type": {"label": "Dokumenttyp", "type": "string"},
            "amount": {"label": "Betrag", "type": "currency"},
        },
        "workflows": {
            "workflow_type": {"label": "Workflow-Typ", "type": "string"},
            "status": {"label": "Status", "type": "string"},
            "started_at": {"label": "Gestartet am", "type": "datetime"},
            "completed_at": {"label": "Abgeschlossen am", "type": "datetime"},
            "duration_ms": {"label": "Dauer (ms)", "type": "number"},
        },
    }

    # Erlaubte Filter-Operatoren
    ALLOWED_OPERATORS = {
        "eq", "ne", "gt", "gte", "lt", "lte",
        "contains", "starts_with", "ends_with",
        "in", "is_null", "is_not_null",
    }

    # Feld-Name-Validierung: nur alphanumerisch und Unterstrich
    _FIELD_NAME_PATTERN = re.compile(r"^[a-zA-Z_][a-zA-Z0-9_]*$")

    # =========================================================================
    # CRUD-Operationen
    # =========================================================================

    async def create_report(
        self,
        db: AsyncSession,
        company_id: "uuid.UUID",
        user_id: "uuid.UUID",
        name: str,
        description: Optional[str],
        data_sources: List[str],
        columns: List[Dict[str, object]],
        filters: Optional[List[Dict[str, object]]] = None,
        grouping: Optional[List[Dict[str, object]]] = None,
        aggregations: Optional[List[Dict[str, object]]] = None,
        chart_config: Optional[Dict[str, object]] = None,
    ) -> AdHocReport:
        """Neuen Ad-Hoc Report erstellen.

        Args:
            db: Datenbank-Session
            company_id: Mandanten-ID
            user_id: Ersteller-ID
            name: Report-Name
            description: Beschreibung (optional)
            data_sources: Liste der Datenquellen-Keys
            columns: Spaltendefinitionen
            filters: Filter (optional)
            grouping: Gruppierung (optional)
            aggregations: Aggregationen (optional)
            chart_config: Chart-Konfiguration (optional)

        Returns:
            Erstellter AdHocReport

        Raises:
            ValueError: Bei ungültigen Datenquellen oder Feldern
        """
        # Datenquellen validieren
        for source in data_sources:
            if source not in self.DATA_SOURCES:
                raise ValueError(
                    f"Ungültige Datenquelle: {source}. "
                    f"Erlaubt: {', '.join(self.DATA_SOURCES.keys())}"
                )

        # Spalten-Felder validieren
        for col in columns:
            source = str(col.get("source", ""))
            field = str(col.get("field", ""))
            self._validate_field(source, field)

        # Filter-Felder validieren
        if filters:
            for f in filters:
                field = str(f.get("field", ""))
                # Feld-Format: source.field oder nur field
                if "." in field:
                    source, field_name = field.split(".", 1)
                    self._validate_field(source, field_name)
                # Operator validieren
                operator = str(f.get("operator", ""))
                if operator and operator not in self.ALLOWED_OPERATORS:
                    raise ValueError(f"Ungültiger Operator: {operator}")

        # Aggregations-Felder validieren
        if aggregations:
            for agg in aggregations:
                agg_type = str(agg.get("type", ""))
                if agg_type and agg_type not in {e.value for e in AggregationType}:
                    raise ValueError(f"Ungültiger Aggregationstyp: {agg_type}")

        report = AdHocReport(
            company_id=company_id,
            created_by=user_id,
            name=name,
            description=description,
            data_sources=data_sources,
            columns=columns,
            filters=filters or [],
            grouping=grouping or [],
            aggregations=aggregations or [],
            chart_config=chart_config,
        )
        db.add(report)
        await db.flush()

        logger.info(
            "adhoc_report_created",
            report_id=str(report.id),
            name=name,
            data_sources=data_sources,
            company_id=str(company_id),
        )
        return report

    async def get_report(
        self,
        db: AsyncSession,
        report_id: "uuid.UUID",
        company_id: "uuid.UUID",
    ) -> Optional[AdHocReport]:
        """Report nach ID laden (mit company_id-Prüfung)."""
        result = await db.execute(
            select(AdHocReport).where(
                and_(
                    AdHocReport.id == report_id,
                    AdHocReport.company_id == company_id,
                )
            )
        )
        return result.scalar_one_or_none()

    async def update_report(
        self,
        db: AsyncSession,
        report_id: "uuid.UUID",
        company_id: "uuid.UUID",
        user_id: "uuid.UUID",
        **kwargs: object,
    ) -> Optional[AdHocReport]:
        """Report aktualisieren (nur Besitzer oder mit Bearbeitungsrecht)."""
        report = await self.get_report(db, report_id, company_id)
        if not report:
            return None

        # Berechtigung prüfen
        if report.created_by != user_id:
            has_edit = await self._user_can_edit(db, report_id, user_id)
            if not has_edit:
                return None

        # Felder validieren falls aktualisiert
        if "data_sources" in kwargs:
            for source in kwargs["data_sources"]:  # type: ignore[union-attr]
                if source not in self.DATA_SOURCES:
                    raise ValueError(f"Ungültige Datenquelle: {source}")

        if "columns" in kwargs:
            for col in kwargs["columns"]:  # type: ignore[union-attr]
                source = str(col.get("source", ""))
                field = str(col.get("field", ""))
                self._validate_field(source, field)

        # Erlaubte Felder aktualisieren
        allowed_fields = {
            "name", "description", "data_sources", "columns", "filters",
            "grouping", "aggregations", "chart_config", "is_public", "is_template",
        }
        for key, value in kwargs.items():
            if key in allowed_fields:
                setattr(report, key, value)

        await db.flush()

        logger.info(
            "adhoc_report_updated",
            report_id=str(report_id),
            updated_fields=list(set(kwargs.keys()) & allowed_fields),
        )
        return report

    async def delete_report(
        self,
        db: AsyncSession,
        report_id: "uuid.UUID",
        company_id: "uuid.UUID",
        user_id: "uuid.UUID",
    ) -> bool:
        """Report löschen (nur Besitzer)."""
        report = await self.get_report(db, report_id, company_id)
        if not report or report.created_by != user_id:
            return False

        await db.delete(report)
        await db.flush()

        logger.info(
            "adhoc_report_deleted",
            report_id=str(report_id),
        )
        return True

    async def list_reports(
        self,
        db: AsyncSession,
        company_id: "uuid.UUID",
        user_id: "uuid.UUID",
        include_shared: bool = True,
        limit: int = 100,
        offset: int = 0,
    ) -> List[AdHocReport]:
        """Verfügbare Reports auflisten (eigene + geteilte + öffentliche)."""
        conditions = [
            AdHocReport.company_id == company_id,
        ]

        # Eigene Reports + öffentliche Reports
        own_or_public = or_(
            AdHocReport.created_by == user_id,
            AdHocReport.is_public == True,  # noqa: E712
        )

        if include_shared:
            # Geteilte Reports als Subquery
            shared_ids_query = (
                select(AdHocReportShare.report_id)
                .where(AdHocReportShare.shared_with_user_id == user_id)
            )
            own_or_public = or_(
                own_or_public,
                AdHocReport.id.in_(shared_ids_query),
            )

        conditions.append(own_or_public)

        result = await db.execute(
            select(AdHocReport)
            .where(and_(*conditions))
            .order_by(AdHocReport.updated_at.desc())
            .limit(limit)
            .offset(offset)
        )
        return list(result.scalars().all())

    # =========================================================================
    # Report-Ausführung
    # =========================================================================

    async def execute_report(
        self,
        db: AsyncSession,
        report_id: "uuid.UUID",
        company_id: "uuid.UUID",
        user_id: "uuid.UUID",
        parameter_overrides: Optional[Dict[str, FilterValue]] = None,
    ) -> Dict[str, object]:
        """Report ausführen und Ergebnisse zurückgeben.

        Args:
            db: Datenbank-Session
            report_id: Report-ID
            company_id: Mandanten-ID
            user_id: Ausführer-ID
            parameter_overrides: Laufzeit-Parameter (optional)

        Returns:
            Dict mit columns, rows, total_rows, execution_time_ms

        Raises:
            ValueError: Bei ungültigem Report oder fehlender Berechtigung
        """
        report = await self.get_report(db, report_id, company_id)
        if not report:
            raise ValueError("Report nicht gefunden oder kein Zugriff")

        start_time = time.monotonic()

        # SQL-Query bauen
        query_text, params = self._build_query(
            report, company_id, parameter_overrides
        )

        # Ausführen mit Timeout (30 Sekunden)
        try:
            result = await db.execute(
                text(f"SET LOCAL statement_timeout = '30000'")
            )
            result = await db.execute(text(query_text), params)
            rows = [dict(row._mapping) for row in result.fetchall()]
        except Exception as e:
            execution_time_ms = int((time.monotonic() - start_time) * 1000)
            # Execution mit Fehler protokollieren
            execution = AdHocReportExecution(
                report_id=report_id,
                company_id=company_id,
                executed_by=user_id,
                row_count=0,
                execution_time_ms=execution_time_ms,
                parameters=parameter_overrides,
                error_message=safe_error_detail(e, "Report"),
            )
            db.add(execution)
            await db.flush()
            raise

        execution_time_ms = int((time.monotonic() - start_time) * 1000)

        # Spalten-Info extrahieren
        column_info = self._extract_column_info(report)

        # Execution protokollieren
        execution = AdHocReportExecution(
            report_id=report_id,
            company_id=company_id,
            executed_by=user_id,
            row_count=len(rows),
            execution_time_ms=execution_time_ms,
            parameters=parameter_overrides,
        )
        db.add(execution)

        # Report-Statistiken aktualisieren
        report.execution_count = (report.execution_count or 0) + 1
        report.last_executed_at = datetime.now(timezone.utc)
        await db.flush()

        logger.info(
            "adhoc_report_executed",
            report_id=str(report_id),
            row_count=len(rows),
            execution_time_ms=execution_time_ms,
        )

        return {
            "columns": column_info,
            "rows": rows,
            "total_rows": len(rows),
            "execution_time_ms": execution_time_ms,
            "execution_id": str(execution.id),
        }

    # =========================================================================
    # Query-Builder (SICHER gegen SQL-Injection)
    # =========================================================================

    def _build_query(
        self,
        report: AdHocReport,
        company_id: "uuid.UUID",
        parameter_overrides: Optional[Dict[str, FilterValue]] = None,
    ) -> Tuple[str, Dict[str, object]]:
        """SQL-Query aus Report-Definition bauen.

        KRITISCH: Nur Whitelist-validierte Feldnamen werden verwendet.
        Alle Werte werden als benannte Parameter gebunden.

        Returns:
            Tuple von (query_string, params_dict)
        """
        params: Dict[str, object] = {"company_id": company_id}
        data_sources = report.data_sources or []
        columns = report.columns or []
        filters = report.filters or []
        grouping = report.grouping or []
        aggregations = report.aggregations or []

        if not data_sources:
            raise ValueError("Report hat keine Datenquellen")

        # --- SELECT ---
        select_parts: List[str] = []
        for col in columns:
            source = str(col.get("source", data_sources[0]))
            field = str(col.get("field", ""))
            alias = str(col.get("alias", ""))

            self._validate_field(source, field)
            table_alias = self._table_alias(source)
            col_ref = f"{table_alias}.{field}"

            if alias and self._FIELD_NAME_PATTERN.match(alias):
                select_parts.append(f"{col_ref} AS {alias}")
            else:
                select_parts.append(col_ref)

        # Aggregationen hinzufügen
        for agg in aggregations:
            field = str(agg.get("field", ""))
            agg_type = str(agg.get("type", "count"))
            alias = str(agg.get("alias", ""))

            if "." in field:
                source, field_name = field.split(".", 1)
            else:
                source = data_sources[0]
                field_name = field

            self._validate_field(source, field_name)
            table_alias = self._table_alias(source)
            col_ref = f"{table_alias}.{field_name}"

            agg_func = self._validate_agg_function(agg_type)
            agg_expr = f"{agg_func}({col_ref})"

            if alias and self._FIELD_NAME_PATTERN.match(alias):
                select_parts.append(f"{agg_expr} AS {alias}")
            else:
                select_parts.append(
                    f"{agg_expr} AS {agg_func}_{field_name}"
                )

        if not select_parts:
            # Fallback: Alle Felder der ersten Datenquelle
            source = data_sources[0]
            table_alias = self._table_alias(source)
            for field in self.AVAILABLE_FIELDS.get(source, []):
                select_parts.append(f"{table_alias}.{field}")

        select_clause = ", ".join(select_parts)

        # --- FROM ---
        primary_source = data_sources[0]
        primary_config = self.DATA_SOURCES[primary_source]
        primary_table = str(primary_config["table"])
        primary_alias = self._table_alias(primary_source)
        from_clause = f"{primary_table} AS {primary_alias}"

        # JOINs für zusätzliche Datenquellen
        join_clauses: List[str] = []
        for i, source in enumerate(data_sources[1:], start=1):
            source_config = self.DATA_SOURCES[source]
            table = str(source_config["table"])
            alias = self._table_alias(source)
            join_key = str(source_config["key"])

            # Standard-Join über company_id
            join_clauses.append(
                f"LEFT JOIN {table} AS {alias} "
                f"ON {alias}.company_id = {primary_alias}.company_id"
            )

        # --- WHERE ---
        where_parts: List[str] = [
            f"{primary_alias}.company_id = :company_id"
        ]

        # Datenquellen-spezifische Filter (z.B. entity_type für suppliers)
        for source in data_sources:
            source_config = self.DATA_SOURCES[source]
            static_filter = source_config.get("filter")
            if static_filter and isinstance(static_filter, dict):
                alias = self._table_alias(source)
                for fk, fv in static_filter.items():
                    if self._FIELD_NAME_PATTERN.match(fk):
                        param_key = f"sf_{source}_{fk}"
                        where_parts.append(f"{alias}.{fk} = :{param_key}")
                        params[param_key] = fv

        # User-Filter
        for idx, f in enumerate(filters):
            field_path = str(f.get("field", ""))
            operator = str(f.get("operator", "eq"))
            value = f.get("value")
            logic = str(f.get("logic", "and")).lower()

            # Override-Wert anwenden
            if parameter_overrides and field_path in parameter_overrides:
                value = parameter_overrides[field_path]

            if "." in field_path:
                source, field_name = field_path.split(".", 1)
            else:
                source = primary_source
                field_name = field_path

            self._validate_field(source, field_name)
            table_alias = self._table_alias(source)
            col_ref = f"{table_alias}.{field_name}"
            param_key = f"f_{idx}"

            condition = self._build_filter_condition(
                col_ref, operator, value, param_key, params
            )
            if condition:
                where_parts.append(condition)

        where_clause = " AND ".join(where_parts)

        # --- GROUP BY ---
        group_by_parts: List[str] = []
        for g in grouping:
            field = str(g.get("field", ""))
            if "." in field:
                source, field_name = field.split(".", 1)
            else:
                source = primary_source
                field_name = field

            self._validate_field(source, field_name)
            table_alias = self._table_alias(source)
            group_by_parts.append(f"{table_alias}.{field_name}")

        group_by_clause = ""
        if group_by_parts:
            group_by_clause = f"GROUP BY {', '.join(group_by_parts)}"

        # --- ORDER BY ---
        order_parts: List[str] = []
        for col in columns:
            sort_dir = str(col.get("sort_direction", "")).upper()
            if sort_dir in ("ASC", "DESC"):
                source = str(col.get("source", primary_source))
                field = str(col.get("field", ""))
                self._validate_field(source, field)
                table_alias = self._table_alias(source)
                order_parts.append(f"{table_alias}.{field} {sort_dir}")

        order_by_clause = ""
        if order_parts:
            order_by_clause = f"ORDER BY {', '.join(order_parts)}"

        # --- Zusammenbauen ---
        query = (
            f"SELECT {select_clause} "
            f"FROM {from_clause} "
            f"{' '.join(join_clauses)} "
            f"WHERE {where_clause} "
            f"{group_by_clause} "
            f"{order_by_clause} "
            f"LIMIT 10000"
        )

        return query.strip(), params

    def _build_filter_condition(
        self,
        col_ref: str,
        operator: str,
        value: object,
        param_key: str,
        params: Dict[str, object],
    ) -> Optional[str]:
        """Einzelne Filter-Bedingung bauen (parametrisiert)."""
        if operator not in self.ALLOWED_OPERATORS:
            return None

        if operator == "eq":
            params[param_key] = value
            return f"{col_ref} = :{param_key}"
        elif operator == "ne":
            params[param_key] = value
            return f"{col_ref} != :{param_key}"
        elif operator == "gt":
            params[param_key] = value
            return f"{col_ref} > :{param_key}"
        elif operator == "gte":
            params[param_key] = value
            return f"{col_ref} >= :{param_key}"
        elif operator == "lt":
            params[param_key] = value
            return f"{col_ref} < :{param_key}"
        elif operator == "lte":
            params[param_key] = value
            return f"{col_ref} <= :{param_key}"
        elif operator == "contains":
            params[param_key] = f"%{value}%"
            return f"{col_ref} ILIKE :{param_key}"
        elif operator == "starts_with":
            params[param_key] = f"{value}%"
            return f"{col_ref} ILIKE :{param_key}"
        elif operator == "ends_with":
            params[param_key] = f"%{value}"
            return f"{col_ref} ILIKE :{param_key}"
        elif operator == "is_null":
            return f"{col_ref} IS NULL"
        elif operator == "is_not_null":
            return f"{col_ref} IS NOT NULL"
        elif operator == "in":
            if isinstance(value, list):
                in_params: List[str] = []
                for i, v in enumerate(value):
                    pk = f"{param_key}_{i}"
                    params[pk] = v
                    in_params.append(f":{pk}")
                return f"{col_ref} IN ({', '.join(in_params)})"
            return None

        return None

    # =========================================================================
    # Validierung
    # =========================================================================

    def _validate_field(self, source: str, field: str) -> None:
        """Feldnamen gegen Whitelist validieren (SQL-Injection-Schutz).

        Raises:
            ValueError: Bei ungültigem Feld
        """
        if not source or source not in self.AVAILABLE_FIELDS:
            raise ValueError(f"Ungültige Datenquelle: {source}")

        if not field or not self._FIELD_NAME_PATTERN.match(field):
            raise ValueError(f"Ungültiger Feldname: {field}")

        allowed = self.AVAILABLE_FIELDS[source]
        if field not in allowed:
            raise ValueError(
                f"Feld '{field}' nicht erlaubt für Datenquelle '{source}'. "
                f"Erlaubt: {', '.join(allowed)}"
            )

    def _validate_agg_function(self, agg_type: str) -> str:
        """Aggregationsfunktion validieren."""
        allowed = {"sum": "SUM", "count": "COUNT", "avg": "AVG", "min": "MIN", "max": "MAX"}
        if agg_type.lower() not in allowed:
            raise ValueError(f"Ungültige Aggregation: {agg_type}")
        return allowed[agg_type.lower()]

    def _table_alias(self, source: str) -> str:
        """Tabellenalias für eine Datenquelle generieren."""
        return f"t_{source}"

    def _extract_column_info(self, report: AdHocReport) -> List[Dict[str, str]]:
        """Spalten-Metadaten für die Antwort extrahieren."""
        columns = report.columns or []
        data_sources = report.data_sources or []
        result: List[Dict[str, str]] = []

        for col in columns:
            source = str(col.get("source", data_sources[0] if data_sources else ""))
            field = str(col.get("field", ""))
            alias = str(col.get("alias", field))
            visible = col.get("visible", True)

            if not visible:
                continue

            metadata = self.FIELD_METADATA.get(source, {}).get(field, {})
            result.append({
                "key": alias or field,
                "field": field,
                "source": source,
                "label": metadata.get("label", field),
                "type": metadata.get("type", "string"),
            })

        return result

    # =========================================================================
    # Export
    # =========================================================================

    async def export_report(
        self,
        db: AsyncSession,
        report_id: "uuid.UUID",
        company_id: "uuid.UUID",
        user_id: "uuid.UUID",
        export_format: AdHocExportFormat,
    ) -> Tuple[bytes, str]:
        """Report exportieren als CSV, Excel oder PDF.

        Returns:
            Tuple von (file_bytes, content_type)
        """
        # Report ausführen
        result = await self.execute_report(
            db, report_id, company_id, user_id
        )
        columns = result["columns"]
        rows = result["rows"]

        if export_format == AdHocExportFormat.CSV:
            return self._export_csv(columns, rows), "text/csv; charset=utf-8"
        elif export_format == AdHocExportFormat.EXCEL:
            return self._export_excel(columns, rows), (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        elif export_format == AdHocExportFormat.PDF:
            return self._export_pdf(columns, rows), "application/pdf"
        else:
            raise ValueError(f"Unbekanntes Export-Format: {export_format}")

    def _export_csv(
        self,
        columns: List[Dict[str, str]],
        rows: List[Dict[str, object]],
    ) -> bytes:
        """CSV-Export generieren."""
        output = io.StringIO()
        writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_MINIMAL)

        # Header
        headers = [str(col.get("label", col.get("key", ""))) for col in columns]
        writer.writerow(headers)

        # Datenzeilen
        for row in rows:
            row_data = [
                str(row.get(col.get("key", col.get("field", "")), ""))
                for col in columns
            ]
            writer.writerow(row_data)

        return output.getvalue().encode("utf-8-sig")

    def _export_excel(
        self,
        columns: List[Dict[str, str]],
        rows: List[Dict[str, object]],
    ) -> bytes:
        """Excel-Export generieren."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            logger.warning("openpyxl_not_installed", fallback="csv")
            return self._export_csv(columns, rows)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"

        # Header-Stil
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")

        # Header schreiben
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col.get("label", col.get("key", ""))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Daten schreiben
        for row_idx, row in enumerate(rows, 2):
            for col_idx, col in enumerate(columns, 1):
                key = col.get("key", col.get("field", ""))
                value = row.get(key, "")
                ws.cell(row=row_idx, column=col_idx).value = value

        # Spaltenbreite anpassen
        for col_idx, col in enumerate(columns, 1):
            max_len = len(str(col.get("label", "")))
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 50)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def _export_pdf(
        self,
        columns: List[Dict[str, str]],
        rows: List[Dict[str, object]],
    ) -> bytes:
        """PDF-Export (Fallback auf CSV wenn reportlab nicht verfügbar)."""
        try:
            from app.services.reports.pdf_export_service import PdfExportService
            from app.services.reports.report_templates import ReportColumn

            pdf_service = PdfExportService()
            pdf_columns = [
                ReportColumn(
                    key=col.get("key", col.get("field", "")),
                    label=col.get("label", ""),
                    format_type=col.get("type", "text"),
                )
                for col in columns
            ]

            import asyncio
            loop = asyncio.get_event_loop()
            if loop.is_running():
                # Synchron aufrufen innerhalb laufendem Loop
                import concurrent.futures
                with concurrent.futures.ThreadPoolExecutor() as pool:
                    pdf_bytes = pool.submit(
                        asyncio.run,
                        pdf_service.generate_report_pdf(
                            title="Ad-Hoc Report",
                            subtitle=f"Erstellt am {datetime.now(timezone.utc).strftime('%d.%m.%Y %H:%M')} UTC",
                            columns=pdf_columns,
                            data=[dict(r) for r in rows],
                            company_name="Ablage-System",
                        )
                    ).result()
            else:
                pdf_bytes = asyncio.run(
                    pdf_service.generate_report_pdf(
                        title="Ad-Hoc Report",
                        subtitle=f"Erstellt am {datetime.now(timezone.utc).strftime('%d.%m.%Y %H:%M')} UTC",
                        columns=pdf_columns,
                        data=[dict(r) for r in rows],
                        company_name="Ablage-System",
                    )
                )
            return pdf_bytes
        except Exception as e:
            logger.warning(
                "pdf_export_fallback_csv",
                error=str(e),
            )
            return self._export_csv(columns, rows)

    # =========================================================================
    # Sharing
    # =========================================================================

    async def share_report(
        self,
        db: AsyncSession,
        report_id: "uuid.UUID",
        owner_id: "uuid.UUID",
        share_with: "uuid.UUID",
        can_edit: bool = False,
    ) -> Optional[AdHocReportShare]:
        """Report mit einem Kollegen teilen."""
        # Besitzer prüfen
        result = await db.execute(
            select(AdHocReport).where(
                and_(
                    AdHocReport.id == report_id,
                    AdHocReport.created_by == owner_id,
                )
            )
        )
        report = result.scalar_one_or_none()
        if not report:
            return None

        # Existierende Freigabe prüfen
        existing = await db.execute(
            select(AdHocReportShare).where(
                and_(
                    AdHocReportShare.report_id == report_id,
                    AdHocReportShare.shared_with_user_id == share_with,
                )
            )
        )
        if existing.scalar_one_or_none():
            raise ValueError("Report ist bereits mit diesem Benutzer geteilt")

        share = AdHocReportShare(
            report_id=report_id,
            shared_with_user_id=share_with,
            can_edit=can_edit,
        )
        db.add(share)
        await db.flush()

        logger.info(
            "adhoc_report_shared",
            report_id=str(report_id),
            shared_with=str(share_with),
            can_edit=can_edit,
        )
        return share

    async def remove_share(
        self,
        db: AsyncSession,
        report_id: "uuid.UUID",
        owner_id: "uuid.UUID",
        share_id: "uuid.UUID",
    ) -> bool:
        """Freigabe entfernen."""
        # Besitzer prüfen
        result = await db.execute(
            select(AdHocReport).where(
                and_(
                    AdHocReport.id == report_id,
                    AdHocReport.created_by == owner_id,
                )
            )
        )
        if not result.scalar_one_or_none():
            return False

        del_result = await db.execute(
            delete(AdHocReportShare).where(
                and_(
                    AdHocReportShare.id == share_id,
                    AdHocReportShare.report_id == report_id,
                )
            )
        )
        await db.flush()
        return del_result.rowcount > 0

    async def _user_can_edit(
        self,
        db: AsyncSession,
        report_id: "uuid.UUID",
        user_id: "uuid.UUID",
    ) -> bool:
        """Prüfen ob ein Benutzer den Report bearbeiten darf."""
        result = await db.execute(
            select(AdHocReportShare).where(
                and_(
                    AdHocReportShare.report_id == report_id,
                    AdHocReportShare.shared_with_user_id == user_id,
                    AdHocReportShare.can_edit == True,  # noqa: E712
                )
            )
        )
        return result.scalar_one_or_none() is not None

    # =========================================================================
    # Datenquellen-Metadaten
    # =========================================================================

    async def get_available_fields(
        self, data_source: str
    ) -> List[Dict[str, str]]:
        """Verfügbare Felder für eine Datenquelle zurückgeben."""
        if data_source not in self.AVAILABLE_FIELDS:
            raise ValueError(f"Ungültige Datenquelle: {data_source}")

        fields = self.AVAILABLE_FIELDS[data_source]
        metadata = self.FIELD_METADATA.get(data_source, {})

        return [
            {
                "field": field,
                "label": metadata.get(field, {}).get("label", field),
                "type": metadata.get(field, {}).get("type", "string"),
                "source": data_source,
            }
            for field in fields
        ]

    def get_data_sources(self) -> List[Dict[str, str]]:
        """Verfügbare Datenquellen zurückgeben."""
        return [
            {
                "key": key,
                "label": str(config.get("label", key)),
                "table": str(config.get("table", "")),
            }
            for key, config in self.DATA_SOURCES.items()
        ]

    # =========================================================================
    # Scheduling
    # =========================================================================

    async def schedule_report(
        self,
        db: AsyncSession,
        report_id: "uuid.UUID",
        company_id: "uuid.UUID",
        frequency: ReportScheduleFrequency,
        export_format: AdHocExportFormat,
        recipients: List[str],
        time_of_day: str = "08:00",
        day_of_week: Optional[int] = None,
        day_of_month: Optional[int] = None,
    ) -> ReportSchedule:
        """Report-Zeitplan einrichten (automatischer Versand per Email).

        Args:
            report_id: Report-ID
            company_id: Mandanten-ID
            frequency: Frequenz (daily/weekly/monthly/quarterly)
            export_format: Export-Format
            recipients: E-Mail-Empfänger
            time_of_day: Uhrzeit im Format HH:MM
            day_of_week: Wochentag (0=Montag) für WEEKLY
            day_of_month: Monatstag (1-28) für MONTHLY/QUARTERLY

        Returns:
            Erstellter ReportSchedule
        """
        # Report prüfen
        report = await self.get_report(db, report_id, company_id)
        if not report:
            raise ValueError("Report nicht gefunden")

        # Zeitformat validieren
        if not re.match(r"^\d{2}:\d{2}$", time_of_day):
            raise ValueError("Ungültige Uhrzeit. Format: HH:MM")

        hour, minute = map(int, time_of_day.split(":"))
        if not (0 <= hour <= 23 and 0 <= minute <= 59):
            raise ValueError("Ungültige Uhrzeit")

        # Nächste Ausführung berechnen
        next_run = self._calculate_next_run(
            frequency, time_of_day, day_of_week, day_of_month
        )

        schedule = ReportSchedule(
            report_id=report_id,
            company_id=company_id,
            frequency=frequency.value,
            day_of_week=day_of_week,
            day_of_month=day_of_month,
            time_of_day=time_of_day,
            export_format=export_format.value,
            recipients=recipients,
            next_run_at=next_run,
        )
        db.add(schedule)
        await db.flush()

        logger.info(
            "adhoc_report_scheduled",
            report_id=str(report_id),
            frequency=frequency.value,
            next_run_at=next_run.isoformat(),
        )
        return schedule

    async def update_schedule(
        self,
        db: AsyncSession,
        schedule_id: "uuid.UUID",
        company_id: "uuid.UUID",
        **kwargs: object,
    ) -> Optional[ReportSchedule]:
        """Zeitplan aktualisieren."""
        result = await db.execute(
            select(ReportSchedule).where(
                and_(
                    ReportSchedule.id == schedule_id,
                    ReportSchedule.company_id == company_id,
                )
            )
        )
        schedule = result.scalar_one_or_none()
        if not schedule:
            return None

        allowed = {
            "frequency", "day_of_week", "day_of_month", "time_of_day",
            "export_format", "recipients", "is_active",
        }
        for key, value in kwargs.items():
            if key in allowed:
                setattr(schedule, key, value)

        # Nächste Ausführung neu berechnen wenn aktiv
        if schedule.is_active:
            schedule.next_run_at = self._calculate_next_run(
                ReportScheduleFrequency(schedule.frequency),
                schedule.time_of_day,
                schedule.day_of_week,
                schedule.day_of_month,
            )

        await db.flush()
        return schedule

    async def delete_schedule(
        self,
        db: AsyncSession,
        schedule_id: "uuid.UUID",
        company_id: "uuid.UUID",
    ) -> bool:
        """Zeitplan löschen."""
        result = await db.execute(
            delete(ReportSchedule).where(
                and_(
                    ReportSchedule.id == schedule_id,
                    ReportSchedule.company_id == company_id,
                )
            )
        )
        await db.flush()
        return result.rowcount > 0

    async def list_schedules(
        self,
        db: AsyncSession,
        company_id: "uuid.UUID",
    ) -> List[ReportSchedule]:
        """Alle Zeitpläne eines Mandanten auflisten."""
        result = await db.execute(
            select(ReportSchedule)
            .where(ReportSchedule.company_id == company_id)
            .order_by(ReportSchedule.next_run_at)
        )
        return list(result.scalars().all())

    async def get_due_schedules(
        self,
        db: AsyncSession,
        limit: int = 50,
    ) -> List[ReportSchedule]:
        """Fällige Zeitpläne laden (für Celery Beat)."""
        now = datetime.now(timezone.utc)
        result = await db.execute(
            select(ReportSchedule)
            .where(
                and_(
                    ReportSchedule.is_active == True,  # noqa: E712
                    ReportSchedule.next_run_at <= now,
                )
            )
            .order_by(ReportSchedule.next_run_at)
            .limit(limit)
        )
        return list(result.scalars().all())

    async def mark_schedule_as_sent(
        self,
        db: AsyncSession,
        schedule: ReportSchedule,
    ) -> None:
        """Markiert einen Zeitplan als gesendet und berechnet den nächsten Zeitpunkt."""
        now = datetime.now(timezone.utc)
        schedule.last_run_at = now
        schedule.next_run_at = self._calculate_next_run(
            ReportScheduleFrequency(schedule.frequency),
            schedule.time_of_day,
            schedule.day_of_week,
            schedule.day_of_month,
        )
        await db.flush()

    def _calculate_next_run(
        self,
        frequency: ReportScheduleFrequency,
        time_of_day: str,
        day_of_week: Optional[int] = None,
        day_of_month: Optional[int] = None,
    ) -> datetime:
        """Nächsten Ausführungszeitpunkt berechnen."""
        now = datetime.now(timezone.utc)
        hour, minute = map(int, time_of_day.split(":"))

        if frequency == ReportScheduleFrequency.DAILY:
            next_run = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
            if next_run <= now:
                next_run += timedelta(days=1)

        elif frequency == ReportScheduleFrequency.WEEKLY:
            target_day = day_of_week if day_of_week is not None else 0
            days_ahead = target_day - now.weekday()
            if days_ahead < 0:
                days_ahead += 7
            next_run = now + timedelta(days=days_ahead)
            next_run = next_run.replace(hour=hour, minute=minute, second=0, microsecond=0)
            if next_run <= now:
                next_run += timedelta(weeks=1)

        elif frequency == ReportScheduleFrequency.MONTHLY:
            target_day = day_of_month if day_of_month is not None else 1
            next_run = now.replace(day=min(target_day, 28), hour=hour, minute=minute, second=0, microsecond=0)
            if next_run <= now:
                if now.month == 12:
                    next_run = next_run.replace(year=now.year + 1, month=1)
                else:
                    next_run = next_run.replace(month=now.month + 1)

        elif frequency == ReportScheduleFrequency.QUARTERLY:
            target_day = day_of_month if day_of_month is not None else 1
            quarter_months = [1, 4, 7, 10]
            current_quarter_month = quarter_months[(now.month - 1) // 3]
            next_quarter_month = quarter_months[((now.month - 1) // 3 + 1) % 4]
            next_year = now.year if next_quarter_month > current_quarter_month else now.year + 1
            next_run = datetime(
                next_year, next_quarter_month, min(target_day, 28),
                hour, minute, 0, tzinfo=timezone.utc
            )
            if next_run <= now:
                idx = ((now.month - 1) // 3 + 2) % 4
                next_year = now.year if quarter_months[idx] > now.month else now.year + 1
                next_run = datetime(
                    next_year, quarter_months[idx], min(target_day, 28),
                    hour, minute, 0, tzinfo=timezone.utc
                )
        else:
            next_run = now + timedelta(days=1)

        return next_run


# Singleton-Pattern
_adhoc_report_service: Optional[AdHocReportService] = None


def get_adhoc_report_service() -> AdHocReportService:
    """Singleton-Instanz des AdHocReportService."""
    global _adhoc_report_service
    if _adhoc_report_service is None:
        _adhoc_report_service = AdHocReportService()
    return _adhoc_report_service
